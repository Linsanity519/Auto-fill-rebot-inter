"""价格面板配置：读 Excel + 跑之前的校验。

⚠ 只服务 mode: price_panel。别的配置类型走不到这里。

一行 = 一个单元。活动是**本批共用一个**（界面上选「本次新建活动」或「挂到已有活动」），
不在 Excel 的单元表里逐行填 —— 和资源位投放同一套规矩。

和资源位投放最大的不同：**套餐排列直接按面板写**。
Excel 里两列「面板1套餐」「面板2套餐」，各自是逗号分隔的 SKU 名，顺序就是
页面上从左到右的顺序；「sku选择」= 两列的并集，机器人自己去勾。
这样人不用再想「选了几个、分隔线该在第几个」——那本来就是页面的实现细节。

值只有两处来，优先级从低到高：

  策略中心  跟着「策略」走的：定向、投放设置、内容限制、算法策略、搭售入口，
            以及方案组「SKU选择 + 面板套餐 + 套餐排序」（含搭售方案名），
            能按单元名称关键词切方案。
  Excel     跟着「这个单元」走的：名字、时间、人群、生效渠道、收银台类型、
            选中类型、赠单片，以及创意层那几列。

⚠ 没有第三个地方。「投放配置」页上只有一件事：选 Excel。
  三个地方找一个值太难了 —— 排查「这个值哪来的」会变成翻三处。

搭售的具体内容（每个 SKU 配什么 pid、买赠什么商品）不在上面任何一处，
在**一份单独的 PID 映射表 Excel** 里，策略中心只存「用哪一套（方案名）」。
一百多个 pid 在界面上逐个填不现实，而且 pid 本来就是从后台一批批抄下来的。
"""
from __future__ import annotations

import logging
import re

from . import wizard_schema as W
from . import wizard_strategy as S
from .datasource import load_table
from .filler import split_multi
from .wizard_data import DataError

log = logging.getLogger(__name__)

UNIT_NAME = "单元名称"
# 三段面板，顺序就是页面上从左到右。面板个数由「后面几段是不是空的」推出来。
PANEL_KEYS = ["面板1套餐", "面板2套餐", "隐藏sku面板套餐"]
PANEL1, PANEL2 = PANEL_KEYS[0], PANEL_KEYS[1]


# ---------------------------------------------------------------- cfg 读取
def unit_fields(cfg: dict) -> list[dict]:
    """Excel「单元」表要出的列。

    excel_from_unit 里的字段定义在 unit_common（页面控件那份），这里只按名字借过来 ——
    同一个字段的 label / 选项 / 必填 / 级联只该有一处。
    它 reveals 出来的子字段自动跟过来（选中类型 → 面板1/2选中套餐）。
    """
    out = [dict(f) for f in (cfg.get("unit_fields") or [])]
    names = [str(n) for n in (cfg.get("excel_from_unit") or [])]
    if not names:
        return out
    # ⚠ 策略中心管的字段不能被顺带拉进 Excel。「内容设置」和「赠单片」都挂在
    #   「收银台类型」的 reveals 底下，但前者归策略中心、后者进 Excel ——
    #   只按「父字段在不在清单里」拉的话，会把内容设置那一串也拉进来。
    owned = set(W.strategy_names(cfg))
    keep = set(names)
    for f in W.flatten(W.unit_fields(cfg, position(cfg))):
        when = f.get("_when")
        if f["name"] in owned:
            continue
        if f["name"] in keep or (when and when[0] in keep):
            keep.add(f["name"])
            out.append(f)
    return out


def sku_list(cfg: dict) -> list[str]:
    return [str(s) for s in (cfg.get("skus") or [])]


# ---------------------------------------------------------------- 活动层
def activity_spec(cfg: dict) -> dict:
    return dict(cfg.get("activity") or {})


def activity_sheet(cfg: dict) -> str:
    return str(activity_spec(cfg).get("sheet") or "活动")


def activity_fields(cfg: dict) -> list[dict]:
    """「活动」sheet 的列。只有「本次新建活动」时才用得上。"""
    return W.flatten(activity_spec(cfg).get("fields") or [])


# ---------------------------------------------------------------- 按 SKU 展开的列
# Excel 里「<SKU>·<字段>」那批列**不枚举 26 个 SKU**，由策略中心那一组方案推出来：
#   哪些 SKU 上面板 ← 三段面板（面板1/面板2/隐藏sku面板）
#   哪些 SKU 搭售   ← 买赠SKU / 0元购SKU（界面上是套餐卡片右边那个小角标）
# 人在策略中心点完套餐和角标，再「生成模板」，Excel 里就正好是这次要填的那几列。
# 改了方案要重新生成一次模板。
#
# 两种层：
#   单元层  sku_unit_fields —— 上了面板就要在单元页上填（目前只有异形SKU 那两项）
#   创意层  creative        —— 创意页上填，字段随搭售类型变
SKU_SEP = "·"


def creative_spec(cfg: dict) -> dict:
    return dict(cfg.get("creative") or {})


def _tie_parts(tie: str) -> list[str]:
    """「买赠+0元购」→ ['买赠', '0元购']；「无」/空 → []。"""
    return [p for p in str(tie or "").replace("＋", "+").split("+")
            if p.strip() and p.strip() != "无"]


def tie_of(values: dict, sku: str) -> str:
    """这个 SKU 的搭售类型，从策略里的两个清单推出来。

    ⚠ 页面上「搭售类型」是六选一的单选，后台却是两个互不相干的字段
      （sale_strategy / add_type）。这里按后台那样存两个清单，两个都命中
      就是「买赠+0元购」—— 不然还得防着几个桶互相打架。
    """
    bits = []
    if sku in split_multi(str((values or {}).get("买赠SKU", "") or "")):
        bits.append("买赠")
    if sku in split_multi(str((values or {}).get("0元购SKU", "") or "")):
        bits.append("0元购")
    return "+".join(bits) or "无"


def panel_skus_of(values: dict) -> list[str]:
    """三段面板里出现过的 SKU，保持书写顺序、去重。"""
    out = []
    for seg in panels_for(values):
        for s in seg:
            if s not in out:
                out.append(s)
    return out


def sku_plan(cfg: dict, payload: dict | None) -> list[tuple[str, str]]:
    """[(SKU, 搭售类型)]，按面板顺序。

    ⚠ 取的是这一组**所有方案的并集**：生成模板时还不知道每个单元叫什么名字，
      「按单元名称匹配」可能命中任何一套，少出一列就等于那个单元填不了。
    """
    grp = S.group_of(payload, "panel")
    order, tie = [], {}
    for name in (grp.get("schemes") or {}):
        vals = S.scheme_values(payload, "panel", name)
        for sku in panel_skus_of(vals):
            if sku not in order:
                order.append(sku)
            t = tie_of(vals, sku)
            # 并集：同一个 SKU 在 A 方案不搭售、B 方案买赠，列要按「买赠」出
            if _tie_parts(t) and not _tie_parts(tie.get(sku, "")):
                tie[sku] = t
            tie.setdefault(sku, t)
    return [(s, tie.get(s) or "无") for s in order]


def _fields_for_sku(cfg: dict, sku: str, tie: str) -> list[tuple[str, dict]]:
    """[(层, 字段)]。层 = 单元层 / 创意层，只用来给模板配色和分组说明。"""
    out = [("单元层", dict(f))
           for f in W.flatten((cfg.get("sku_unit_fields") or {}).get(sku) or [])]
    if sku in (cfg.get("sku_map_skip") or []):
        # 异形SKU 页面上没有「搭售类型」这一档，by_tie 那两组对它不适用
        tie = "无"
    spec = creative_spec(cfg)
    out += [("创意层", dict(f)) for f in (spec.get("sku_common") or [])]
    for part in _tie_parts(tie):
        out += [("创意层", dict(f))
                for f in W.flatten((spec.get("by_tie") or {}).get(part) or [])]
    out += [("创意层", dict(f))
            for f in W.flatten((spec.get("by_sku") or {}).get(sku) or [])]
    return out


def sku_columns(cfg: dict, payload: dict | None) -> list[dict]:
    """单元 sheet 里按 SKU 展开的列，列名「<SKU>·<字段>」。

    面板级那两列（切换按钮）只有真配了面板2 才出 —— 页面上「面板个数=1个」时
    根本没有这两个输入框。
    """
    cols: list[dict] = []
    for sku, tie in sku_plan(cfg, payload):
        for layer, f in _fields_for_sku(cfg, sku, tie):
            col = dict(f)
            col["name"] = f"{sku}{SKU_SEP}{f['name']}"
            col["_sku"] = {"sku": sku, "tie": tie, "field": f["name"], "layer": layer}
            # _when 里的父字段也要带上 SKU 前缀，不然级联指到别的 SKU 头上
            when = f.get("_when")
            if when:
                col["_when"] = (f"{sku}{SKU_SEP}{when[0]}", list(when[1]))
            cols.append(col)
    if _has_panel2(payload):
        for f in (creative_spec(cfg).get("panel_fields") or []):
            col = dict(f)
            col["_sku"] = {"sku": "", "tie": "", "field": f["name"], "layer": "创意层"}
            cols.append(col)
    return cols


def _has_panel2(payload: dict | None) -> bool:
    grp = S.group_of(payload, "panel")
    for name in (grp.get("schemes") or {}):
        vals = S.scheme_values(payload, "panel", name)
        if split_multi(str(vals.get(PANEL2, "") or "")) or \
           split_multi(str(vals.get(PANEL_KEYS[2], "") or "")):
            return True
    return False


def tie_options(cfg: dict) -> list[str]:
    return [str(s) for s in (cfg.get("搭售类型选项") or ["无"])]


def sku_map(cfg: dict, data: dict, values: dict, sku: str) -> dict:
    """这个 SKU 这次该怎么配搭售。

    搭不搭售、搭哪种 —— 一律看策略里的「买赠SKU / 0元购SKU」（界面上那个角标）。
    搭什么、用哪个 pid —— 去 PID 映射表里按「PID映射方案」捞，pid 按这个单元的
    「生效平台」取一组（页面上那一栏是多选）。

    异形SKU 例外：页面上它没有「搭售类型」这一档，就固定一对 pid + 搭售商品，
    而且这两个值一个单元一个样，所以从 Excel 那两列里取（sku_unit_fields）。
    """
    if sku in (cfg.get("sku_map_skip") or []):
        out = {k.split(SKU_SEP, 1)[1]: str(v).strip()
               for k, v in (values or {}).items() if k.startswith(f"{sku}{SKU_SEP}")}
        if out.get("价格面板pid"):
            out["pid清单"] = [out["价格面板pid"]]
        out["搭售商品"] = out.get("搭售商品ID", "")
        return out

    tie = tie_of(values, sku)
    plan = str((values or {}).get("PID映射方案", "")).strip()
    item = dict(((data.get("pid_map") or {}).get(plan) or {}).get(sku) or {}) if plan else {}
    # ⚠ 搭售类型不从映射表读 —— 那张表只回答「搭什么」，不回答「搭不搭」。
    #   两边都能说同一件事的话，不一致时谁赢没人讲得清。
    item["搭售类型"] = tie
    if tie == "无":
        return item
    pids, missing = pids_for_platforms(cfg, item, (values or {}).get("生效平台", ""))
    item["pid清单"] = pids
    item["缺平台"] = missing
    return item


def pid_sheet(cfg: dict) -> str:
    return str((cfg.get("pid_sheet") or {}).get("sheet") or "PID映射")


def pid_columns(cfg: dict) -> list[str]:
    return [str(c) for c in (cfg.get("pid_sheet") or {}).get("columns") or []]


def load_pid_map(path: str, cfg: dict) -> dict:
    """读 PID 映射表 → {方案名: {SKU: {"pids": {平台: pid}, 搭售类型, 买赠商品类型, ...}}}。

    ⚠ 一个 (方案, SKU) 有**好几行**，一个平台一行 —— 页面上「价格面板pid」是
      多选，一个单元投几个平台就填几个 pid。
    ⚠ 搭售类型 / 买赠商品 这些是「一个 SKU 一份」，不是「一个平台一份」：
      在任意一行填一次就行，这里取第一个非空的。逼人在 9 行里重复填同一个值，
      改的时候一定会漏。
    ⚠ 方案名留空的行归到 ""（空方案名）下，等于「这份表只有一套」。
    """
    try:
        rows = load_table(path, pid_sheet(cfg))
    except Exception:
        rows = load_table(path, None)

    shared = [c for c in pid_columns(cfg) if c not in ("方案名", "SKU", "平台", "价格面板pid")]
    out: dict[str, dict] = {}
    for row in rows:
        sku = str(row.get("SKU", "") or "").strip()
        if not sku:
            continue
        plan = str(row.get("方案名", "") or "").strip()
        item = out.setdefault(plan, {}).setdefault(sku, {"pids": {}})
        plat = str(row.get("平台", "") or "").strip()
        pid = str(row.get("价格面板pid", "") or "").strip()
        if pid:
            item["pids"][plat] = pid
        for c in shared:
            v = str(row.get(c, "") or "").strip()
            if v and not item.get(c):
                item[c] = v
    return out


def platform_alias(cfg: dict) -> dict:
    """价格表里的平台叫法 → 页面「生效平台」的叫法。"""
    return {str(k): str(v) for k, v in (cfg.get("pid_platform_alias") or {}).items()}


def pids_for_platforms(cfg: dict, item: dict, effective: str) -> tuple[list, list]:
    """这个 SKU 在这个单元该填哪几个 pid，外加「哪些平台没配到」。

    按单元的「生效平台」取：投几个平台就填几个 pid（页面上那一栏是多选）。
    """
    alias = platform_alias(cfg)
    want = [x.strip() for x in str(effective or "").replace("，", ",").split(",") if x.strip()]
    pids, missing = [], []
    for page_plat in want:
        hit = [pid for sheet_plat, pid in (item.get("pids") or {}).items()
               if alias.get(sheet_plat, sheet_plat) == page_plat]
        if hit:
            for pid in hit:
                if pid not in pids:
                    pids.append(pid)
        else:
            missing.append(page_plat)
    return pids, missing


def position(cfg: dict) -> str:
    return str(cfg.get("position") or W.position_names(cfg)[0])


def strategy_for(cfg: dict, payload: dict, unit_name: str) -> dict:
    """这个单元该套用的策略值。按单元名称走关键词方案就靠这一步。"""
    return S.resolve(cfg, payload, position(cfg), unit_name)


def values_for(cfg: dict, data: dict, unit: dict) -> dict:
    """这个单元最终要往页面上填的一整份值。取值只走这一个口子。

    优先级（后面的盖前面的）：策略中心 → Excel 这一行。
    「面板个数」不在任何一处配，是从三段面板推出来的。
    """
    name = str(unit["header"].get(UNIT_NAME, ""))
    vals = dict(strategy_for(cfg, data.get("strategy") or {}, name))
    # Excel 这一行填了什么就盖什么（活动/名字/时间，以及借过来的那几个页面字段）
    for k, v in (unit["header"] or {}).items():
        if str(v).strip():
            vals[k] = str(v).strip()
    # 面板个数不配，由三段推出来
    vals["面板个数"] = panel_count_of(panels_for(vals))
    return vals


def panels_for(values: dict) -> list[list[str]]:
    """三段各放哪些 SKU（顺序就是页面上从左到右）。

    返回 [面板1, 面板2, 隐藏sku面板]，末尾的空段**不去掉** ——
    调用方靠「后面几段是不是空的」去推面板个数。
    """
    return [split_multi(str(values.get(k, "") or "")) for k in PANEL_KEYS]


def panel_count_of(panels: list) -> str:
    """由三段推出页面上「面板个数」该选什么。

    ⚠ 这一项不让人配：它和「哪几段有东西」是同一件事，分开配就能配出
      「面板个数=1个 但面板2 排了三个 SKU」这种页面上不存在的状态。
    """
    if panels[2]:
        return "2个+查看更多"
    return "2个" if panels[1] else "1个"


# ---------------------------------------------------------------- 读数据
def _cell_images(data_file: str) -> dict[tuple[int, str], str]:
    """把贴在「单元」表格子里的图片抽成文件，返回 {(行号, 列名): 本地路径}。

    图片列有三种填法，都得支持（和资源位投放一个样，别让人为了这个再另存一遍）：
      · 本地路径      直接用
      · 贴在格子里    就是这里抽出来的
      · http(s) 网址  真去传的时候再下（见 pp_creative.upload）

    ⚠ 抽图要单独开一次 workbook（openpyxl 读 _images 的限制），所以复用
      wizard_data 里那份，别再写一遍。
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    from .wizard_data import _extract_images

    try:
        by_sheet = _extract_images(data_file)
    except Exception:                      # 抽不出图不该拦住整个流程
        log.warning("读单元格图片失败，忽略", exc_info=True)
        return {}
    imgs = by_sheet.get("单元") or (by_sheet.get(next(iter(by_sheet), "")) or {})
    if not imgs:
        return {}

    try:
        wb = load_workbook(data_file, read_only=True)
        ws = wb["单元"] if "单元" in wb.sheetnames else wb.worksheets[0]
        headers = {get_column_letter(i): str(c.value).strip()
                   for i, c in enumerate(next(ws.iter_rows(max_row=1)), 1) if c.value}
        wb.close()
    except Exception:
        log.warning("读表头失败，单元格图片用不上", exc_info=True)
        return {}

    out = {}
    for ref, path in imgs.items():
        col = "".join(ch for ch in ref if ch.isalpha())
        num = "".join(ch for ch in ref if ch.isdigit())
        name = headers.get(col)
        if name and num:
            out[(int(num), name)] = path
            log.info("单元!%s 用了贴在格子里的图片 → %s", ref, name)
    return out


def pid_map_of(cfg: dict, payload: dict | None) -> tuple[dict, str]:
    """按策略里存的路径读 PID 映射表，返回 (map, 出错时的人话)。

    ⚠ 生成模板时也要用它（创意列按搭售类型出），所以从 load() 里抽出来单放。
    """
    pid_path = str(((payload or {}).get("rules") or {}).get("PID映射表", "") or "").strip()
    if not pid_path:
        return {}, ""
    try:
        return load_pid_map(pid_path, cfg), ""
    except Exception as e:                   # 路径写错/文件被占用，别在这儿炸
        return {}, f"PID 映射表读不了（{pid_path}）：{e}"


def load(data_file: str, cfg: dict, settings: dict | None = None) -> dict:
    """读 Excel → {"activity": {...}, "units": [...], "strategy": 当前生效的那套策略}。

    sheet 名不强求：优先找叫「单元」的那张，没有就用第一张 ——
    模板是我们自己生成的，但用户常把它另存/复制到别的表里。

    settings 里认 `wizard_activity`（和资源位投放同一个键，界面上是同一行控件）：
      {"existing": bool, "activity_id": str, "activity_type": str}
    """
    settings = settings or {}
    try:
        rows = load_table(data_file, "单元")
    except Exception:
        rows = load_table(data_file, None)

    payload = S.active_payload(cfg)
    pid_map, pid_error = pid_map_of(cfg, payload)

    defaults = {f["name"]: str(f.get("default", "")) for f in unit_fields(cfg)}
    per_sku = sku_columns(cfg, payload)
    for f in per_sku:
        defaults.setdefault(f["name"], str(f.get("default", "")))
    known = set(defaults)
    # 单元层那几列（异形SKU 的 pid / 搭售商品）要进 header —— sku_map 从 values 里读它们
    sku_only = {f["name"] for f in per_sku if (f["_sku"]["layer"] == "创意层")}

    cell_imgs = _cell_images(data_file)

    units = []
    for i, row in enumerate(rows):
        header, creative = {}, {}
        excel_row = i + 2               # +2：表头占一行，Excel 行号从 1 起
        for name in known:
            val = str(row.get(name, "") or "").strip() or defaults.get(name, "")
            # 那一格没写字、但贴了张图 —— 用图
            if not val:
                val = cell_imgs.get((excel_row, name), "")
            (creative if name in sku_only else header)[name] = val
        if not header.get(UNIT_NAME):
            continue                    # 整行没名字 = 空行/说明行，跳过
        units.append({
            "row": excel_row,
            "header": header,
            "creative": creative,
        })

    return {"activity": _activity(data_file, cfg, settings),
            "units": units, "strategy": payload,
            "pid_map": pid_map, "pid_error": pid_error}


def _activity(data_file: str, cfg: dict, settings: dict) -> dict:
    """活动信息。和资源位投放同一套规矩：

      挂到已有活动  → 用界面上填的活动ID，Excel 里没有活动 sheet
      本次新建活动  → 读「活动」sheet 那一行，跑的时候先建活动再建单元

    ⚠ 返回值里有没有「已有活动ID」是 runner 的分叉点，值从哪来它不用管。
    """
    act_cfg = settings.get("wizard_activity") or {}
    if act_cfg.get("existing"):
        aid = str(act_cfg.get("activity_id", "")).strip()
        if not aid:
            raise DataError("选了「挂到已有活动」，但没填活动ID")
        return {"已有活动ID": aid,
                # 单元页 URL 里要带一个活动类型，填 5（测试验收）不影响建单元
                "活动类型ID": str(act_cfg.get("activity_type", "") or "5").strip(),
                "活动名称": str(act_cfg.get("activity_name", "")).strip()}

    sheet = activity_sheet(cfg)
    try:
        rows = load_table(data_file, sheet)
    except Exception:
        rows = []
    rows = [r for r in rows if any(str(v or "").strip() for v in r.values())]
    if not rows:
        raise DataError(
            f"模板里没有「{sheet}」sheet，或者那张表一行没填。"
            "要挂到已有活动的话，在「投放配置」页把活动切到「挂到已有活动」并填活动ID；"
            "要新建活动就重新生成一份带活动 sheet 的模板。")
    if len(rows) > 1:
        raise DataError(f"「{sheet}」sheet 有 {len(rows)} 行，只能填一行 —— 本批单元都挂在这一个活动下")
    return {str(k): str(v or "").strip() for k, v in rows[0].items()}


# ---------------------------------------------------------------- 校验
_TIME_RE = re.compile(r"^\d{4}-\d{2}-\d{2}([ T]\d{2}:\d{2}(:\d{2})?)?$")


def _check_activity(cfg: dict, act: dict) -> list[str]:
    """活动层。本批共用一个活动，所以这里只查一次，不逐行查。"""
    exist = str(act.get("已有活动ID", "")).strip()
    if exist:
        if not exist.isdigit():
            return [f"[活动] 「活动ID」要填数字，现在是「{exist}」"]
        return []
    out = []
    for f in activity_fields(cfg):
        when = f.get("_when")
        if when and not W.when_active(f, act.get(when[0], "")):
            continue
        if f.get("required") and not str(act.get(f["name"], "")).strip():
            out.append(f"[活动] 「{f['name']}」没填")
    for col in ("活动开始时间", "活动结束时间"):
        v = str(act.get(col, "")).strip()
        if v and not _TIME_RE.match(v):
            out.append(f"[活动] 「{col}」格式不对：「{v}」，要写成 2026-09-24 10:00:00")
    return out


def validate(cfg: dict, data: dict) -> list[str]:
    """跑之前把能在本地看出来的问题一次说完，别等填到一半才炸。

    ⚠ 值是**逐个单元**算的（同一批单元可能按名字命中不同的方案），
      所以校验也得逐个单元来，不能拿一份全局的值糊弄过去。
    """
    issues: list[str] = []
    payload = data.get("strategy") or {}
    pos = position(cfg)
    all_skus = set(sku_list(cfg))
    units = data.get("units") or []

    if data.get("pid_error"):
        issues.append(data["pid_error"])
    if not units:
        issues.append("Excel 里一个单元都没读到（「单元名称」这一列是空的？）")
    issues += _check_activity(cfg, data.get("activity") or {})

    seen_names = {}

    for u in units:
        tag = f"第{u['row']}行"
        h = u["header"]
        name = h.get(UNIT_NAME, "")
        vals = values_for(cfg, data, u)
        panels = panels_for(vals)
        flat = [x for seg in panels for x in seg]

        # 名字没命中任何关键词、组里又没兜底 —— 说人话，别报成「某字段没配」
        issues += [f"{tag}{x}" for x in S.unmatched_hint(cfg, payload, name)]
        issues += _check_required(cfg, vals, pos, tag)

        if len(name) > 24:
            issues.append(f"{tag}「{UNIT_NAME}」超过 24 字（页面限制），现在 {len(name)} 字")
        if name in seen_names:
            issues.append(f"{tag}和第{seen_names[name]}行的单元名称一样，确认不是复制漏改？")
        else:
            seen_names[name] = u["row"]

        for col in ("投放开始时间", "投放结束时间"):
            v = str(h.get(col, "")).strip()
            if v and not _TIME_RE.match(v):
                issues.append(f"{tag}「{col}」格式不对：{v}（要 2026-09-24 10:00:00）")

        if not panels[0]:
            issues.append(f"{tag}面板1 一个 SKU 都没有 —— 策略中心「SKU选择 + 面板套餐 + "
                          f"套餐排序」那一组里，这个单元命中的方案没给面板1")
        if panels[2] and not panels[1]:
            issues.append(f"{tag}选了「隐藏sku面板」但面板2 是空的 —— "
                          f"页面上没有「跳过面板2」这种排法")

        bad = [x for x in flat if x not in all_skus]
        if bad:
            issues.append(f"{tag}这些不是页面上的 SKU：{'、'.join(bad)}"
                          f"（注意是「超大年度套餐」不是「超大年度会员」）")
        dup = [x for x in set(flat) if flat.count(x) > 1]
        if dup:
            issues.append(f"{tag}同一个 SKU 落在了两个面板里：{'、'.join(dup)}")

        issues += _check_sku_map(cfg, data, vals, list(dict.fromkeys(flat)), tag)

        # 「选中类型 = 指定套餐」时挑的那个套餐，得真的在这个单元的对应面板里
        if str(vals.get("选中类型", "")).strip() == "指定套餐":
            for key, seg in (("面板1选中套餐", panels[0]), ("面板2选中套餐", panels[1])):
                want = str(vals.get(key, "")).strip()
                if want and want not in seg:
                    issues.append(f"{tag}「{key}」选的是「{want}」，但它不在那个面板里")

    return issues


def _check_required(cfg: dict, vals: dict, pos: str, tag: str) -> list[str]:
    """该有值而没值的必填字段（不管它是策略中心供的还是投放配置页供的）。

    ⚠ 只查「这个单元这次真的会出现」的字段：条件字段（内容类型只在
      内容设置=指定 时才有）没触发就不该拦，否则界面上根本没那一项、
      却一直提示没填。触发判断统一走 wizard_schema.when_active。
    """
    out = []
    excel_names = {f["name"] for f in unit_fields(cfg)}
    for f in W.flatten(W.unit_fields(cfg, pos)):
        if f.get("scope") in ("manual", "derived") or not f.get("required"):
            continue
        when = f.get("_when")
        if when and not W.when_active(f, vals.get(when[0], "")):
            continue
        if str(vals.get(f["name"], "")).strip():
            continue
        where = "Excel 里这一行" if f["name"] in excel_names else "策略中心"
        out.append(f"{tag}{where}的「{f['name']}」没填")
    return out


def _check_sku_map(cfg: dict, data: dict, vals: dict, skus: list, tag: str) -> list[str]:
    """这个单元用到的每个 SKU，搭售配全了没有。

    ⚠ 只检查这个单元真正用到的 SKU —— 映射表里 25 个 SKU 大多数用不上，
      全查一遍会刷出一堆和本次投放无关的红字。
    """
    out = []
    allowed = tie_options(cfg)
    plan = str(vals.get("PID映射方案", "")).strip()
    pid_map = data.get("pid_map") or {}
    if plan and plan not in pid_map:
        return [f"{tag}策略里选的 PID 映射方案是「{plan}」，但 PID 映射表里没有这个方案名"
                f"（表里有：{'、'.join(pid_map) or '一套都没有'}）"]

    for sku in skus:
        m = sku_map(cfg, data, vals, sku)
        if sku in (cfg.get("sku_map_skip") or []):
            for col in ("价格面板pid", "搭售商品ID"):
                if not m.get(col):
                    out.append(f"{tag}用到了「{sku}」，但 Excel 里「{sku}{SKU_SEP}{col}」这一列没填")
            continue

        tie = m.get("搭售类型", "") or "无"
        if tie == "无":
            continue
        if tie not in allowed:
            out.append(f"{tag}「{sku}」的搭售类型是「{tie}」，本期只适配 {'、'.join(allowed)}")
            continue
        if not plan:
            out.append(f"{tag}「{sku}」在策略里标了「{tie}」，但没选 PID 映射方案 —— "
                       f"pid 和商品无处可取")
            continue
        if not m or not any(k in m for k in ("pids", "买赠商品类型", "组合价格")):
            out.append(f"{tag}PID 映射表的方案「{plan}」里没有「{sku}」这一行，"
                       f"但策略里给它标了「{tie}」")
            continue
        where = f"PID 映射表（方案「{plan}」/ {sku}）"
        if "买赠" in tie:
            if not m.get("pid清单"):
                out.append(f"{tag}{where}的搭售类型是{tie}，但这个单元的生效平台"
                           f"（{vals.get('生效平台', '?')}）一个 pid 都没配到")
            elif m.get("缺平台"):
                out.append(f"{tag}{where}：这几个平台在映射表里没有 pid —— "
                           f"{'、'.join(m['缺平台'])}，它们不会带搭售")
            for col in ("买赠商品类型", "买赠商品ID"):
                if not m.get(col):
                    out.append(f"{tag}{where}的搭售类型是{tie}，但「{col}」没填")
        if "0元购" in tie:
            pairs = combine_pairs(m.get("组合价格", ""))
            if not pairs:
                out.append(f"{tag}{where}的搭售类型是{tie}，但「组合价格」没填"
                           f"（写成「单会员价格pid:加购商品id」，多组用逗号分隔）")
            for pr in pairs:
                if len(pr) != 2 or not all(pr):
                    out.append(f"{tag}{where}的组合价格「{':'.join(pr)}」不成对，"
                               f"要写成「单会员价格pid:加购商品id」")
    return out


def combine_pairs(value: str) -> list[list[str]]:
    """「11439:100,11440:101」→ [['11439','100'], ['11440','101']]。"""
    out = []
    for chunk in split_multi(str(value or "")):
        parts = [x.strip() for x in re.split(r"[:：]", chunk)]
        out.append(parts)
    return out
