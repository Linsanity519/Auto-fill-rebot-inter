"""原生商广的 Excel 模板：一张平表 + 一页填写说明。

⚠ 只服务 mode: ad_native。

和资源位投放的模板不一样，这里不用「单元名称留空 = 上一个单元的又一条创意」
那套写法 —— 运营手里的素材表本来就是一行一条素材、带个内容/剧集列，
直接按「内容」列聚类更贴近他们已有的表，少一道人工整理。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import ad_data as D
from . import ad_prep as P
from .paths import user_path

REQ_FILL = PatternFill("solid", fgColor="FFF2CC")    # 必填 浅黄
OPT_FILL = PatternFill("solid", fgColor="F2F2F2")    # 选填 浅灰
KEY_FILL = PatternFill("solid", fgColor="DDEBF7")    # 聚类键 浅蓝


def build(cfg: dict, prep: dict | None = None, out_name: str | None = None) -> str:
    if prep is None:
        prep = P.load(cfg)

    wb = Workbook()
    ws = wb.active
    ws.title = "素材"

    kcol = D.key_column(cfg)
    cols = D.columns(cfg)
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c["name"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = (KEY_FILL if c["name"] == kcol
                     else REQ_FILL if c.get("required") else OPT_FILL)
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(42, len(c["name"]) * 3))
        note = c.get("_note")
        if note:
            cell.comment = Comment(note, "配置助手")
    ws.freeze_panes = "A2"

    _doc_sheet(wb, cfg, prep)

    name = out_name or f"{cfg['name']}_模板"
    out = user_path("data", f"{name}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)


def _doc_sheet(wb, cfg: dict, prep: dict):
    doc = wb.create_sheet("填写说明")
    for col, w in (("A", 24), ("B", 26), ("C", 80)):
        doc.column_dimensions[col].width = w

    kcol = D.key_column(cfg)
    cap = D.max_creatives(cfg)
    g = cfg.get("grouping") or {}

    rows = [
        ("配置类型", cfg["name"], cfg.get("description", "")),
        ("", "", ""),
        ("怎么填", "", "① 一行 = 一条创意。封面和素材标题一一对应，不会跨行合并。"),
        ("", "", f"②「{kcol}」相同的行会自动合成一个单元，单元名按 "
                 f"{g.get('name_template', '')} 拼。"),
        ("", "", f"③ 一个单元最多 {cap} 条创意；同一个{kcol}超过 {cap} 行会自动拆成多个单元，"
                 f"名字按 {g.get('overflow_template', '')} 加后缀，序号从 0 顺延（_0、_1、_2…）。"),
        ("", "", f"④ 同一个{kcol}的行不用挨着写，程序按首次出现的顺序归组。"),
        ("", "", "⑤ 同一个 avid 可以写多行 —— 同一条视频配不同封面/不同标题，"
                 "会在页面上建成多条独立创意。"),
        ("", "", "⑥「素材描述」留空会自动取「内容」列的值 —— 这两列平时就是同一个剧集名，"
                 "只有剧集名超过 10 字（页面硬上限）时才需要在描述列单独填个短的。"),
        ("", "", "⑦ 计划名称、转化目标、出价、投放时间、人群这些在界面「准备」页填，不进这张表。"),
        ("", "", ""),
        ("颜色", "", "蓝=聚类键　黄=必填　灰=选填"),
        ("封面列", "", "选填。留空 = 用稿件自己的原始封面；"
                       "要换就填本地图片路径（如 D:\\素材\\a.png），或直接把图片贴进单元格。"),
        ("", "", ""),
        ("■ 列说明", "", ""),
    ]
    for c in D.columns(cfg):
        rows.append((f"　{c['name']}", "必填" if c.get("required") else "选填",
                     c.get("_note", "")))

    rows += [("", "", ""), ("■ 准备阶段当前的值", "", "改这些请回界面「准备」页")]
    for name, val in P.summary(cfg, prep):
        rows.append((f"　{name}", val, ""))

    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if str(v).startswith("■") or str(v) in ("怎么填", "颜色"):
                cell.font = Font(bold=True)
