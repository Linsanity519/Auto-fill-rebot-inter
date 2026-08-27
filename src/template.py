"""从表单配置生成 Excel 模板。

表头直接由 config/forms/<名字>.yaml 推导，配置改了重新生成即可，
不会出现"模板列名和配置对不上"这种低级故障。

⚠ 必须是可直接 import 调用的函数：打包成 exe 后 sys.executable 就是 exe 本身，
  用 subprocess 去跑脚本只会再开一个界面。
"""
import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .datasource import header_field_names, item_field_names
from . import xlsx_kit as X
from .paths import user_path

# 颜色统一在 src/xlsx_kit.py 里定义（原来六份模板各声明了一遍同样的十六进制值）
REQ_FILL = X.FILLS["req"]    # 必填 - 浅黄
OPT_FILL = X.FILLS["opt"]    # 选填 - 浅灰
KEY_FILL = X.FILLS["key"]    # 分组 - 浅蓝


def collect_meta(form_cfg: dict) -> dict:
    """{字段名: {required, options, scope}}"""
    meta = {}

    def add(f, scope):
        meta.setdefault(f["name"], {
            "required": f.get("required", False),
            "options": f.get("options"),
            "match": f.get("match"),
            # 复选组是多选，值形如 "iPhone,Android"
            "multi": f.get("type") == "checkbox_sync",
            "scope": scope,
        })

    for f in form_cfg["fields"]:
        add(f, "主表")
        for subs in (f.get("reveals") or {}).values():
            for s in subs:
                add(s, "主表(条件)")

    list_cfg = form_cfg.get("list") or {}
    groups = list(list_cfg.get("variants", {}).values()) or [list_cfg.get("fields", [])]
    for g in groups:
        for f in g:
            add(f, "明细")
    return meta


def build(form_name: str) -> str:
    """生成模板，返回文件路径。"""
    path = user_path("config", "forms", f"{form_name}.yaml")
    cfg = yaml.safe_load(path.read_text(encoding="utf-8"))

    meta = collect_meta(cfg)
    # 「分组」只在有明细行（list）时才有意义——它的作用就是把多行并成一条配置。
    # 一行一条的表单（如 DMP人群新建）加这一列只会让人以为漏填了什么。
    grouped = bool(cfg.get("list"))
    cols = (["分组"] if grouped else []) + header_field_names(cfg) + item_field_names(cfg.get("list"))

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        m = meta.get(c)
        cell.fill = KEY_FILL if c == "分组" else (REQ_FILL if m and m["required"] else OPT_FILL)
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(c) * 2.2)

    # 固定选项的列做成 Excel 下拉，从源头堵住填错
    for i, c in enumerate(cols, 1):
        m = meta.get(c) or {}
        opts = m.get("options")
        col = get_column_letter(i)

        # ⚠ 多选列不能做单选下拉：Excel 的 list 校验只认单个选项，
        #   会把「iPhone,Android」这种合法值判为非法，直接把人卡死。
        #   改成批注提示 + 预填一个示例。
        if m.get("multi"):
            ws.cell(row=1, column=i).comment = Comment(
                f"可多选，用逗号分隔\n例：{opts[0]},{opts[1]}\n\n可选值：{'、'.join(opts)}"
                if opts and len(opts) > 1 else "可多选，用逗号分隔",
                "配置助手",
            )
            continue

        if not opts or m.get("match") == "contains":
            continue
        formula = '"' + ",".join(opts) + '"'
        if len(formula) > 255:      # Excel 内联列表长度上限
            continue
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
        dv.error = f"只能填：{'/'.join(opts)}"
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}1000")

    ws.freeze_panes = "B2"
    _write_doc_sheet(wb, cfg, cols, meta, grouped)

    return X.save(wb, f"{form_name}_模板.xlsx")


def _write_doc_sheet(wb, cfg, cols, meta, grouped=True):
    doc = wb.create_sheet("填写说明")
    for col, w in (("A", 22), ("B", 12), ("C", 90)):
        doc.column_dimensions[col].width = w

    if grouped:
        how = [
            ("怎么填", "", "一条配置有几个明细项就写几行，这几行的「分组」填同一个值。"),
            ("", "", "主表字段（人群名称、平台、优先级等）只需在该组第一行填，后续行留空。"),
            ("", "", "黄色列 = 必填，灰色列 = 选填，蓝色「分组」列 = 归组用，不会填进页面。"),
        ]
    else:
        how = [
            ("怎么填", "", "一行 = 一条配置，从第 2 行开始往下写，有多少条写多少行。"),
            ("", "", "黄色列 = 必填，灰色列 = 选填。选填列留空就用页面上的默认值。"),
        ]

    rows = [
        ("表单", cfg["name"], cfg.get("description", "")),
        ("", "", ""),
        *how,
        ("", "", ""),
        ("字段", "是否必填", "可选值 / 说明"),
    ]
    head_row = len(rows)

    for c in (cols[1:] if grouped else cols):
        m = meta.get(c, {})
        opts = m.get("options")
        note = "、".join(opts) if opts else "自由填写"
        if m.get("match") == "contains":
            note = "填 ID（如 35697），页面会自动匹配到对应项"
        elif m.get("multi"):
            eg = ",".join(opts[:2]) if opts and len(opts) > 1 else (opts[0] if opts else "")
            note = f"【可多选】用英文逗号分隔，例：{eg}　　可选值：{'、'.join(opts or [])}"
        rows.append((c, "必填" if m.get("required") else "选填", note))

    variants = (cfg.get("list") or {}).get("variants")
    if variants:
        key = cfg["list"]["variants_by"]
        rows += [("", "", ""), (f"⚠ 按「{key}」不同，要填的明细列不一样", "", "")]
        for v, fs in variants.items():
            rows.append((f"  {key}={v}", "", "只填这几列：" + "、".join(f["name"] for f in fs)))

    reveal_rows = []
    for f in cfg["fields"]:
        for val, subs in (f.get("reveals") or {}).items():
            reveal_rows.append(
                (f"  {f['name']}={val}", "", "才需要填：" + "、".join(s["name"] for s in subs)))
    if reveal_rows:
        rows += [("", "", ""), ("⚠ 选不同值，要填的列也不一样", "", "")] + reveal_rows

    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if r == head_row or str(v).startswith("⚠"):
                cell.font = Font(bold=True)
