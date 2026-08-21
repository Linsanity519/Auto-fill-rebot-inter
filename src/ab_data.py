"""读「指定实验ID 续期」的清单文件。

新增文件，只被 AbRunner 的 id_list 模式用到 —— datasource / wizard_data / dmp_data
一行不动。日期解析复用 src/dmp_date.py，两边对日期的宽容度保持一致。

清单支持 .xlsx / .xlsm / .csv，列名认这几种写法：
  实验ID   ← 实验id / ID / id / 实验编号 / testId / test_id
  延期至   ← 到期日期 / 实验到期日期 / 续期至 / 延期日期 / 目标日期（留空 = 延到最晚）
  实验名称 ← 名称（选填，只用于核对，不参与定位）
"""
from __future__ import annotations

import csv
import logging
from pathlib import Path

from .dmp_date import parse_date

log = logging.getLogger(__name__)

ID_ALIASES = ("实验id", "实验ID", "id", "ID", "实验编号", "testid", "test_id")
DATE_ALIASES = ("延期至", "到期日期", "实验到期日期", "续期至", "延期日期", "目标日期", "延长至")
NAME_ALIASES = ("实验名称", "名称", "实验名")


class AbDataError(Exception):
    pass


def _pick(headers: list[str], aliases) -> str | None:
    """在表头里找一个能对上的列名，大小写和空格都忽略。"""
    norm = {h.strip().lower().replace(" ", ""): h for h in headers if h}
    for a in aliases:
        hit = norm.get(a.strip().lower().replace(" ", ""))
        if hit:
            return hit
    return None


def _read_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            rows = []
            for i, rec in enumerate(csv.DictReader(fh), 2):
                rec["_row"] = i
                rows.append(rec)
            return rows

    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [(str(ws.cell(1, i).value).strip() if ws.cell(1, i).value is not None else "")
               for i in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec, empty = {}, True
        for i, h in enumerate(headers, 1):
            if not h:
                continue
            v = ws.cell(r, i).value
            rec[h] = v
            if v is not None and str(v).strip():
                empty = False
        if not empty:
            rec["_row"] = r
            rows.append(rec)
    return rows


def _clean_id(value) -> str:
    """Excel 会把纯数字 ID 读成 15863.0，这里统一成字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load(path: str) -> list[dict]:
    """读清单，返回 [{id, name, date, date_raw, row, issues}]。

    只做「文件层面」的校验（ID 空、日期认不出、ID 重复）。
    「这个 ID 在页面上搜不搜得到」要连页面才知道，放在 AbRunner.preview 里查。
    """
    p = Path(path)
    if not p.exists():
        raise AbDataError(f"实验清单文件不存在：{p}")
    if p.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
        raise AbDataError(f"实验清单只认 .xlsx / .xlsm / .csv，当前是：{p.suffix}")

    rows = _read_rows(p)
    if not rows:
        raise AbDataError(f"{p.name} 里没有数据行")

    headers = [h for h in rows[0] if h != "_row"]
    id_col = _pick(headers, ID_ALIASES)
    if not id_col:
        raise AbDataError(
            f"{p.name} 里找不到「实验ID」列（表头现在是：{headers}）。"
            f"点「生成 Excel 模板」拿标准表头。")
    date_col = _pick(headers, DATE_ALIASES)
    name_col = _pick(headers, NAME_ALIASES)

    out, seen = [], {}
    for rec in rows:
        row_no = rec.get("_row")
        cid = _clean_id(rec.get(id_col))
        raw = rec.get(date_col) if date_col else None
        # ⚠ 只取第一段是为了容忍「2026-11-11 00:00:00」这种带时间的写法。
        #   CSV 的空单元格是 ""（不是 None），"".split() 是空列表，不兜底会 IndexError。
        raw_s = "" if raw is None else ((str(raw).strip().split() or [""])[0])

        issues = []
        if not cid:
            issues.append(f"第{row_no}行：实验ID 为空")
        elif cid in seen:
            issues.append(f"第{row_no}行：实验ID「{cid}」和第{seen[cid]}行重复")
        else:
            seen[cid] = row_no

        d = parse_date(raw)
        if raw_s and d is None:
            issues.append(f"第{row_no}行：延期至「{raw_s}」认不出来，请写成 2026-11-11 这种格式")

        out.append({
            "id": cid,
            "name": _clean_id(rec.get(name_col)) if name_col else "",
            "date": d,
            "date_raw": raw_s,
            "row": row_no,
            "issues": issues,
        })
    return out
