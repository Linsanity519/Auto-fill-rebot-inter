"""方向 C 的入口：pywebview 窗口 + 本地 HTML（assets/webui/）。

阶段2：三步面板真正接上 Runner。WebUI(BaseUI) 通过 window.evaluate_js 把
log/progress/confirm/ask_continue/finished 推给前端；Runner 本身在后台线程跑，
和 gui.py 的 GuiUI + 队列轮询是同一个模型，只是"轮询"换成了"推送"。

⚠ Api 复用同一个 runner 实例贯穿"载入并检查→ 开始配置"，不像 gui.py.on_start()
  那样重新 _make_runner 一个新实例。这是有意的改动，不是疏忽：
  DmpRunner.run() 只有在 self._targets（预检时缓存的候选人群）还在同一个实例上
  才会用它筛出"预检时选中的那些"；gui.py 每次开跑都新建实例，_targets 早丢了，
  所以 GUI 版这里实际上是"重新扫一遍范围内所有人群"而不是"只跑预检选中的"——
  命令行版 main.py --cli 因为 preview()/run() 用的是同一个实例，反而没有这个问题。
  这里选择复用实例，行为对齐命令行版（更接近直觉：核对页看到啥，跑的就是啥）。

和 gui.py 的关系：这是默认界面；gui.py 留作 --tk 备用选项，出问题时退回去用。
"""
import contextlib
import json
import logging
import os
import queue
import subprocess
import sys
import threading
import time
from pathlib import Path

import webview
import yaml

from . import chrome, formcfg, notify, registry, update, usage
from . import settings as settings_defaults
from .paths import app_dir, resource, user_path
from .ui import BaseUI, Stopped

log = logging.getLogger(__name__)

ROOT = app_dir()
FORMS_DIR = user_path("config", "forms")

# pywebview 6.x 把 OPEN_DIALOG 改叫 FileDialog.OPEN，老名字每次弹框都往日志里
# 刷一条 deprecation 警告。取新的、没有就退回老的，两个版本都能跑。
OPEN_DIALOG = getattr(getattr(webview, "FileDialog", None), "OPEN", None)
if OPEN_DIALOG is None:
    OPEN_DIALOG = webview.OPEN_DIALOG


def _load_settings() -> dict:
    """和 gui.py._load_settings 同样的逻辑：相对路径锚定到 exe 所在目录。"""
    path = user_path("config", "settings.yaml")
    if not path.exists():
        raise FileNotFoundError(
            f"找不到配置文件：{path}\n请确认 config 文件夹和程序放在同一目录下。")
    s = settings_defaults.apply_defaults(yaml.safe_load(path.read_text(encoding="utf-8")))
    for key in ("state_file", "result_file", "log_file", "screenshot_dir", "data_file"):
        v = s.get(key)
        if v and not Path(v).is_absolute():
            s[key] = str(ROOT / v)
    return s


def _json_safe(obj):
    """过一遍 json 往返，防止意外的非 JSON 类型（比如 openpyxl 读出的日期）
    把 evaluate_js/Api 返回值的桥接调用崩掉——宁可显示成字符串，也不能整个调用失败。
    """
    return json.loads(json.dumps(obj, ensure_ascii=False, default=str))


# 选文件对话框的筛选器。
# ⚠ pywebview 对格式很挑：描述部分必须匹配 [\w ]+，斜杠、顿号一概不行
#   （见 webview/util.py 的 parse_file_type）。原来写的是「Excel / CSV (...)」，
#   那个斜杠让 create_file_dialog 直接抛 ValueError，异常又被吞掉，
#   表现就是「点浏览没任何反应」。改这里之前先自测：
#     python -c "import webview.util as u; u.parse_file_type('你的筛选器')"
FILE_TYPES = ("Excel 和 CSV (*.xlsx;*.xlsm;*.csv)", "所有文件 (*.*)")


class WebUI(BaseUI):
    """Runner 通过它和前端通信，全部走 window.evaluate_js 推。

    confirm()/ask_continue() 会阻塞 Runner 所在的后台线程，直到前端弹窗
    点了按钮、Api.answer() 把值塞进 self.answer——和 gui.py 的 GuiUI 是同一个
    队列阻塞模型，只是队列另一头从"tkinter 主循环轮询"换成了"JS 事件回调"。

    ⚠ 存 window 的属性名必须带下划线（_window，不能叫 window）：pywebview 生成
      window.pywebview.api 桥接对象时会反射遍历 js_api（也就是 Api 实例）上
      所有不带下划线的属性，递归下钻。Api.ui 是这里的 WebUI 实例，如果这个类
      也用不带下划线的 window 存 pywebview 的 Window 对象，遍历会顺着
      api.ui.window 一路钻进 .NET WinForms 的 AccessibilityObject/ActiveControl
      这些 COM 属性——这些属性在 pythonnet 包装下会互相递归引用，一直递归到
      Python 的递归上限，而且每次异常都要展开调用栈，代价不小。实测这就是
      窗口一点就"未响应"的真正原因（一开始怀疑是外部安全软件在探测，排查后
      发现纯粹是这里的命名问题，跟外部软件无关）。
    """

    def __init__(self, window):
        self._window = window
        self.answer = queue.Queue(maxsize=1)
        self.stop_flag = threading.Event()
        self.pause_flag = threading.Event()
        # 「逐条确认」模式下等人点按钮的总时长。埋点要用它把「机器在干活」和
        # 「机器在等人」分开——不然一次跑了 40 分钟里有 35 分钟是在等确认，
        # 却全算成机器代劳，这个数就不实了。
        self.wait_seconds = 0.0
        # 每条各花了多久。六个 Runner 都在每条做完后调 progress()，所以在这里掐表
        # 就能拿到逐条耗时，一个 Runner 都不用动。存的是「净耗时」——同样扣掉等人。
        self.item_seconds = []
        self._tick = None
        self._tick_wait = 0.0

    def _push(self, fn, *args):
        try:
            payload = json.dumps(list(args), ensure_ascii=False, default=str)
            self._window.evaluate_js(f"window.app.{fn}.apply(null, {payload})")
        except Exception:
            log.warning("推送前端失败：%s", fn, exc_info=True)

    def log(self, msg, level="info"):
        self._push("onLog", msg, level)

    def progress(self, done, total, stats):
        now = time.monotonic()
        if self._tick is not None and done:
            # 两次 progress 之间就是一条的耗时，减去这中间等人的那部分
            net = (now - self._tick) - (self.wait_seconds - self._tick_wait)
            if net >= 0:
                self.item_seconds.append(net)
        self._tick, self._tick_wait = now, self.wait_seconds
        self._push("onProgress", done, total, dict(stats))

    def confirm(self, label, summary):
        self._push("onConfirm", label, summary)
        with self._waiting():
            while True:
                try:
                    return self.answer.get(timeout=0.3)
                except queue.Empty:
                    if self.stop_flag.is_set():
                        return "stop"

    def ask_continue(self, error):
        self._push("onAskContinue", error)
        with self._waiting():
            while True:
                try:
                    return self.answer.get(timeout=0.3)
                except queue.Empty:
                    if self.stop_flag.is_set():
                        return False

    @contextlib.contextmanager
    def _waiting(self):
        """这段时间是在等人，不是机器在干活。"""
        t0 = time.monotonic()
        try:
            yield
        finally:
            self.wait_seconds += time.monotonic() - t0

    def checkpoint(self):
        if self.stop_flag.is_set():
            raise Stopped()
        while self.pause_flag.is_set():
            if self.stop_flag.is_set():
                raise Stopped()
            threading.Event().wait(0.2)

    def finished(self, title, body, ok):
        notify.beep(ok=ok)
        self._push("onFinished", title, body, ok)


def _prep_field(f: dict) -> dict:
    """一个字段发给界面的样子。

    ⚠ 把 _when 归一成 when：从 unit_common 借来的字段带的是 _when（reveals 展开出来的），
      prep_fields 里自己写的是 when。界面只认一个名字，不然级联会失灵 ——
      表现是「选中类型=默认，底下那两个套餐下拉还杵在那儿」。
    """
    out = dict(f, kind=_prep_kind(f))
    if not out.get("when") and out.get("_when"):
        out["when"] = list(out["_when"])
    return out


def _prep_kind(f: dict) -> str:
    """这一项在界面上长什么样：segmented / select / number / file / text。

    ⚠ 不能直接拿 type 当长相。原生商广的 prep_fields 里 type 写的就是长相
      （segmented/select/number）；价格面板配置的字段是从 unit_common 借来的，
      那边的 type 是**填写方式**（pp_radio / pp_checkbox…，给 pp_filler 用的）。
      只认 type 的话，后者所有字段都会渲染成纯文本框，下拉和分段全没了。
    """
    t = str(f.get("type") or "")
    if t in ("segmented", "select", "number", "file", "text"):
        return t                      # 自己就写的是长相（原生商广）
    if t == "pp_number":
        return "number"
    opts = f.get("options") or []
    if "checkbox" in t or "multiselect" in t:
        return "text"                 # 多选：逗号分隔的文本，选项写在提示里
    if opts:
        return "segmented" if len(opts) <= 3 else "select"
    return "text"


class Api:
    """暴露给前端 JS 的方法（window.pywebview.api.xxx()，前端按 Promise 用）。

    ⚠ Api 的实例本身就是 pywebview 反射遍历生成桥接对象时的遍历起点，所以这个类
      上任何存了"深处挂着 pywebview Window 对象"的属性都必须带下划线前缀，
      理由见 WebUI 类文档字符串。
    """

    def __init__(self, settings: dict):
        self.settings = settings
        self._window = None          # main() 建完窗口后回填
        self.runner = None          # 载入并检查之后的那个 runner 实例，开跑时复用
        self.ui: WebUI | None = None
        self.preview_rows = []      # 上一次 load_and_check 的 PreviewRow 列表
        self.worker: threading.Thread | None = None
        self.last_results = []      # 上一次 run() 的返回值
        self._run_kept_rows = []    # 上一次 run() 时，传给它的 records 每个位置对应哪个 PreviewRow
        # 埋点用：当前在哪个配置类型上、什么范围、上一次运行的 id（重跑要指回去）
        self.form_name = ""
        self.scope = None
        self.last_run_id = None
        self._syncing = False       # 正在往企微表格写，别并发开第二趟
        self._sync_error = ""       # 上一次上报失败的原因，首页要显示出来
        self._flow_rec = None       # 录制中的 FlowRecorder
        self._flow_browser = None   # 录制期间那个长开的 Browser（不走 with）
        from . import __version__
        self._updater = update.UpdateService(settings, __version__)

    # ---------------- 配置类型 ----------------
    #
    # ⚠ 界面上「这个配置类型有没有 xxx」一律由这里按 yaml 算出来发给前端，
    #   前端不许再写 `mode === "wizard" || mode === "price_panel"` 这种清单。
    #
    #   为什么：那种清单在 Python 和 JS 各存一份，接一个新配置类型要两边都改，
    #   而且漏改是**静默的**（卡片不显示，没有任何报错）。
    #   实际发生过：hasStrategy / hasPrepCard 在 app.js 里按 mode 名写死，
    #   而 Python 这边早就改成看 strategy_groups / prep_fields 了 —— 同一个判断
    #   两套实现。现在只有这一处。
    #
    #   ⚠ 下面这个函数体里**一个 mode 名都不该出现**。要是又忍不住写
    #     `mode == "xxx"`，说明该在 yaml 里补一个声明，不是在这里加分支。
    @staticmethod
    def _caps(cfg: dict) -> dict:
        # ⚠ strategy / prep 一律转调各自模块里那个唯一的判据函数，不在这里重写条件。
        #   重写一遍就等于又多一份会走样的实现（prep 的条件是
        #   prep_fields **或** prep_from_unit，价格面板走的正是后者）。
        from . import ad_prep as P
        from . import wizard_strategy as S

        positions = cfg.get("positions") or {}
        return {
            # 策略中心（配一次全批套用，可建多套方案、按单元名关键词切）
            "strategy": S.has_strategy(cfg),
            # 「准备」页那张共用参数平表
            "prep": P.has_prep(cfg),
            # 要不要勾「本次投哪些资源位」。只有一个资源位的类型没这回事
            "positions": len(positions) > 1,
            # 本批共用一个活动：要么本次新建、要么挂到已有
            "activity": bool(cfg.get("activity") or cfg.get("steps")),
            # 抢占任务清单那张卡（日期/时间段/人数/楼栋，在界面上填不走 Excel）
            "task_list": bool(cfg.get("grab")),
            # 吃不吃 Excel 数据文件。不吃的在 yaml 里写 data_source: none
            "excel": cfg.get("data_source", "excel") != "none",
            # 「批量开关」类型：藏掉数据文件行，露出一个「名称关键词」文本框
            "toggle": bool(cfg.get("toggle")),
            # 自制配置类型（录制生成的工作流）：准备页显示步骤卡
            "flow": bool(cfg.get("flow")),
        }

    # 界面上跟着配置类型变的那几句话。yaml 里 ui: 段可以覆盖，
    # 不写就用下面的默认 —— 新接一个类型不用回来改 JS 的文案。
    @staticmethod
    def _ui_text(cfg: dict, caps: dict) -> dict:
        ui = cfg.get("ui") or {}
        multi = caps["positions"]
        return {
            "deliver_label": ui.get("deliver_label") or ("资源位投放配置" if multi else "投放配置"),
            "deliver_hint": ui.get("deliver_hint") or (
                "选资源位 → 生成模板 → 填好 Excel → 载入并检查 → 跑" if multi
                else "配好策略 → 生成模板 → 填好 Excel → 载入并检查 → 跑"),
            "strategy_hint": ui.get("strategy_hint") or
            "配在这里的字段，模板里就不用逐个单元填了",
            # 跑法：fill=填表（空跑/逐条确认/全自动），grab=抢占（只找不订/开抢）
            "run_kind": ui.get("run_kind") or ("grab" if caps["task_list"] else "fill"),
            # 「批量开关」那几个控件的文案（yaml 的 ui: 段可覆盖）
            "params_label": ui.get("params_label") or "名称关键词",
            "params_placeholder": ui.get("params_placeholder")
            or "一行一个关键词，命中即算。留空 = 整页所有行",
            "strategy_label": ui.get("strategy_label") or "策略",
            "strategy_placeholder": ui.get("strategy_placeholder")
            or "留空 = 当前打开的策略页。跨策略：一行一个，编辑页URL / 路由ID / 业务ID",
            "toggle_hint": ui.get("toggle_hint") or "",
        }

    def list_forms(self) -> list:
        from . import flow_data as FD
        out = []
        for name, cfg in list(formcfg.load_all()) + FD.list_all():
            nav = cfg.get("nav") or {}
            caps = self._caps(cfg)
            out.append({
                "name": name,
                # ⚠ mode 仍然发出去，但前端只拿它做日志/埋点，不许拿它判断界面长什么样。
                #   要判断长什么样，看下面的 caps。
                "mode": cfg.get("mode"),
                "caps": caps,
                "ui": self._ui_text(cfg, caps),
                "scopes": registry.scopes_for(cfg),
                # 侧栏归类。yaml 没写 nav 的（比如新加的配置）落到「其他」组，
                # 名字就用文件名 —— 界面照样能显示，不至于漏掉一整项。
                "group": nav.get("group") or "其他",
                "group_order": nav.get("group_order", 99),
                "label": nav.get("label") or name,
                "order": nav.get("order", 99),
                # 首页没数据时当功能导航用：一句话说清这个配置类型是干嘛的
                "desc": cfg.get("description") or "",
            })
        return out

    # ---------------- 首页 ----------------
    def app_info(self) -> dict:
        """版本号显示在侧栏底部。全项目只有 src/__init__.py 那一个版本号。"""
        from . import __version__
        return {"version": __version__}

    # ---------------- 程序更新 ----------------
    def check_update(self, force: bool = False) -> dict:
        """前端启动时后台调用；失败不会影响任何业务功能。"""
        return self._updater.check(force=bool(force))

    def download_update(self) -> dict:
        return self._updater.download()

    def install_update(self, package: str) -> dict:
        """交棒给独立 EXE。当前窗口随后退出，避免占用待替换的程序文件。

        ⚠ 不信前端传回来的路径：verify_downloaded 会重算摘要并确认它就在
          output/updates 里，否则等于让页面指定「拿哪个文件覆盖程序」。
        """
        kind = self._updater.verify_downloaded(package)
        if not kind:
            return {"ok": False, "error": "更新包校验未通过，请重新下载"}
        helper = ROOT / "配置助手更新器.exe"
        if not helper.is_file():
            return {"ok": False, "error": f"找不到更新器：{helper}"}
        log_path = ROOT / "output" / "update-run.log"
        try:
            subprocess.Popen([
                str(helper), "--pid", str(os.getpid()),
                f"--{kind}", package, "--target", str(ROOT), "--log", str(log_path),
            ], cwd=str(ROOT), creationflags=getattr(subprocess, "CREATE_NEW_PROCESS_GROUP", 0))
            # 让本次 API 返回先送达前端，再销毁窗口；更新器会等本进程真正退出。
            threading.Timer(0.5, self._window.destroy).start()
            return {"ok": True}
        except OSError as e:
            log.exception("启动更新器失败")
            return {"ok": False, "error": f"启动更新器失败：{e}"}

    def refresh_team(self) -> dict:
        """前端首屏渲染完之后调：去 GitHub 拉最新团队快照。

        放在首屏之后而不是之前 —— 先用本地那份把界面点亮，拉到新的再重绘，
        网络慢的时候不至于让人对着空白页等。
        """
        try:
            changed = usage.fetch_team(self.settings)
        except Exception:
            log.warning("团队快照刷新失败（不影响使用）", exc_info=True)
            changed = False
        return {"changed": bool(changed)}

    def usage_summary(self) -> dict:
        """首页那些数。口径见 src/usage.py 的 summarize()。

        ⚠ 只读本地：全团队那份来自缓存（跑完一轮 / 启动时 / 点刷新 才去表格拉）。
          读一次表格要开浏览器、全选、复制，十几秒起步，不能挂在每次进首页上。
        """
        try:
            out = usage.summarize(self.settings)
            out["report"] = self._report_status()
            team = usage.load_team()
            if team:
                # 全团队那份是**随包分发的快照**（config/team.json），不是实时的 ——
                # 前端必须标「截至 X/X」。本机自己的数留在 mine 里，那份永远是新的。
                out["mine"] = {k: out[k] for k in ("totals", "week", "longest",
                                                   "forms", "weeks", "recent")}
                out["people"] = team.get("people", out.get("people"))
                out["totals"] = team.get("totals", out["totals"])
                out["forms"] = team.get("forms", out["forms"])
                out["actives"] = team.get("actives") or []
                out["team_weeks"] = team.get("weeks") or {}
                out["snapshot_at"] = team.get("synced_at", "")
            return out
        except Exception as e:
            log.exception("统计聚合失败")
            return {"error": str(e)}

    def _report_status(self) -> dict:
        """本机的上报状态：有没有欠着的周、上次发出去是什么时候、上次为什么失败。

        ⚠ 这一块是 2026-08-21 补的。在这之前上报失败是**完全静默**的（只往
          run.log 写一行），结果是「有人用了一整天，团队表里一条没有」而谁都不知道。
        """
        try:
            from . import report
            names = [f["name"] for f in self.list_forms()]
            return {
                "on": report.enabled(self.settings),
                "pending": usage.pending_weeks(self.settings, names),
                # 团队那份是随包分发的快照，不是本机同步来的，这里给的是快照的日期
                "snapshot_at": (usage.load_team() or {}).get("synced_at", ""),
                "error": self._sync_error,
                "syncing": bool(self._syncing),
            }
        except Exception:
            log.warning("算上报状态失败", exc_info=True)
            return {}

    def _form_cfg(self, form_name: str) -> dict:
        # 自制配置类型：定义在 config/flows/<名>.json，包成 synthetic cfg 走同一套
        from . import flow_data as FD
        if FD.exists(form_name):
            return FD.synthetic_cfg(FD.load(form_name))
        return formcfg.load(form_name)

    # ---------------- 浏览器 ----------------
    def browser_status(self) -> bool:
        return chrome.is_connected(self.settings["cdp_url"], timeout=0.8)

    def launch_browser(self, form_name: str | None) -> dict:
        try:
            url = None
            if form_name:
                try:
                    url = self._form_cfg(form_name).get("form_url")
                except Exception:
                    pass
            url = url or self.settings.get("start_url")
            msg = chrome.launch(self.settings["cdp_url"], ROOT / ".chrome-profile", url)
            return {"ok": True, "message": msg}
        except Exception as e:
            log.exception("启动浏览器失败")
            return {"ok": False, "message": str(e)}

    # ---------------- 准备：选文件 / 生成模板 ----------------
    def pick_file(self) -> str | None:
        """弹系统选文件框。返回路径；用户取消返回 None。

        ⚠ 筛选器再出问题时退回「不带筛选器」重开一次，而不是静默失败 ——
          能选到文件永远比筛得好看重要。
        """
        start_dir = ROOT / "data"
        directory = str(start_dir if start_dir.exists() else ROOT)
        try:
            result = self._window.create_file_dialog(
                OPEN_DIALOG, directory=directory, file_types=FILE_TYPES)
        except Exception:
            log.exception("按筛选器弹选择框失败，退回不带筛选器再试一次")
            try:
                result = self._window.create_file_dialog(
                    OPEN_DIALOG, directory=directory)
            except Exception:
                log.exception("选择文件失败")
                self._toast("打不开选择文件的窗口，请把路径直接粘进输入框。"
                            "详情见 output/run.log", "error")
                return None
        return result[0] if result else None

    def _toast(self, msg: str, level: str = "info"):
        """往界面的日志抽屉推一行。失败也不能让调用方炸掉。"""
        try:
            self._window.evaluate_js(
                f"window.app && window.app.onLog({json.dumps(msg)}, {json.dumps(level)})")
        except Exception:
            log.warning("推日志到界面失败：%s", msg, exc_info=True)

    def wizard_meta(self, form_name: str) -> dict:
        """「准备」页的资源位卡片 + 策略中心界面要的全部静态信息，一次取完。"""
        from . import wizard_schema as W
        from . import wizard_strategy as S

        cfg = self._form_cfg(form_name)
        # ⚠ 判据是「这个配置类型有没有声明策略中心的分组」，不是「mode == wizard」。
        #   价格面板配置也用同一套策略中心（同一个界面、同一份 config/strategies/），
        #   按 mode 写死的话，每接一个新配置类型都要回来改一次 Python。
        if not S.has_strategy(cfg):
            return {"wizard": False}
        positions = []
        for name in W.position_names(cfg):
            meta = W.position_meta(cfg, name)
            positions.append({
                "name": name,
                "scene": meta.get("scene", ""),
                "system": meta.get("system", ""),
                "real_name": meta.get("real_name", ""),
                "strategy_fields": [f["name"] for f in W.strategy_fields_for(cfg, name)],
            })
        return _json_safe({
            "wizard": True,
            "positions": positions,
            "strategy_fields": S.field_defs_for_ui(cfg),
            "groups": list(W.strategy_groups(cfg).keys()),
            # 方案组（人群 / 内容限制）：界面一组一张卡，用法完全一样
            "scheme_groups": S.group_defs_for_ui(cfg),
        })

    # ---------------- 准备阶段参数（原生商广 / 价格面板配置）----------------
    # ⚠ 和 wizard 的「策略中心」是两码事：那边有方案库/关键词匹配/资源位例外，
    #   这边就是一张「字段名 → 值」的平表，界面上直接铺开填，不另开一页。
    #
    # ⚠ 判据是「yaml 里有没有声明 prep_fields」，不是「mode 等于 ad_native」——
    #   价格面板配置也用这张表（26 个 SKU 各自的搭售类型/pid/买赠商品），
    #   按 mode 写死的话，新加一个用这套表的配置类型就要回来改一次 Python。
    def ad_meta(self, form_name: str) -> dict:
        from . import ad_prep as P

        cfg = self._form_cfg(form_name)
        if not P.has_prep(cfg):
            return {"ad": False}
        return _json_safe({
            "ad": True,
            "fields": [_prep_field(f) for f in P.field_defs(cfg)],
            "values": P.load(cfg),
            "grouping": cfg.get("grouping") or {},
        })

    # ---------------- 预定会议室：抢占任务清单 ----------------
    # ⚠ 和 ad_meta 的区别：那边是「一张字段→值的平表，整批共用」，这边是一个**列表**，
    #   每条任务各自的日期/时段/人数都不一样，所以另走一套读写口。
    def meeting_meta(self, form_name: str) -> dict:
        from . import meeting_data as MD

        cfg = self._form_cfg(form_name)
        if cfg.get("mode") != "meeting_reserve":
            return {"meeting": False}
        doc = MD.load(cfg)
        return _json_safe({
            "meeting": True,
            "buildings": cfg.get("buildings") or [],
            "default_task": MD.DEFAULT_TASK,
            "weekday_names": MD.WEEKDAY_NAMES,
            "rule_text": MD.RULE_TEXT,
            "open_time": str((cfg.get("grab") or {}).get("open_time", "10:00")),
            "tasks": doc["tasks"],
            "issues": [{"index": i + 1, "items": MD.validate(t)}
                       for i, t in enumerate(doc["tasks"])],
        })

    def meeting_save(self, form_name: str, tasks: list) -> dict:
        from . import meeting_data as MD

        try:
            cfg = self._form_cfg(form_name)
            path = MD.save(cfg, {"tasks": tasks or []})
            doc = MD.load(cfg)
            return _json_safe({
                "ok": True, "path": path, "tasks": doc["tasks"],
                "issues": [{"index": i + 1, "items": MD.validate(t)}
                           for i, t in enumerate(doc["tasks"])],
            })
        except Exception as e:
            log.exception("存抢占任务失败")
            return {"ok": False, "error": str(e)}

    def prep_save(self, form_name: str, values: dict) -> dict:
        from . import ad_prep as P
        try:
            cfg = self._form_cfg(form_name)
            path = P.save(cfg, values)
            return {"ok": True, "path": path, "values": _json_safe(P.load(cfg)),
                    "issues": P.validate(cfg, values)}
        except Exception as e:
            log.exception("存准备参数失败")
            return {"ok": False, "error": str(e)}

    # ---------------- 策略中心 ----------------
    def strategy_get(self, form_name: str) -> dict:
        from . import wizard_strategy as S
        try:
            cfg = self._form_cfg(form_name)
            return {"ok": True, "doc": _json_safe(S.load(cfg)), "path": str(S.path_for(cfg))}
        except Exception as e:
            log.exception("读策略失败")
            return {"ok": False, "error": str(e)}

    def strategy_save(self, form_name: str, doc: dict) -> dict:
        from . import wizard_strategy as S
        try:
            cfg = self._form_cfg(form_name)
            path = S.save(cfg, doc)
            return {"ok": True, "path": path, "doc": _json_safe(S.load(cfg))}
        except Exception as e:
            log.exception("存策略失败")
            return {"ok": False, "error": str(e)}

    def make_template(self, form_name: str, scope: str | None, positions: list | None = None,
                      options: dict | None = None) -> dict:
        try:
            cfg = self._form_cfg(form_name)
            if cfg.get("mode") == "flow":
                # 自制配置类型：按 flow 的 data.columns 出一张空表
                from . import flow_data as FD
                cols = FD.columns(cfg.get("_flow") or {})
                if not cols:
                    return {"ok": False, "error": "这个工作流没有绑任何 Excel 列，不用模板"}
                path = self._flow_template(form_name, cols)
                usage.record(self.settings, "template_made", form=form_name, entry="webui")
                return {"ok": True, "path": str(path)}
            if cfg.get("mode") == "ad_native":
                from . import ad_template as AT
                path = AT.build(cfg)
                usage.record(self.settings, "template_made", form=form_name, entry="webui")
                return {"ok": True, "path": str(path)}
            if cfg.get("mode") == "wizard":
                from . import wizard_template as WT
                if not positions:
                    return {"ok": False, "error": "至少勾一个资源位"}
                opts = options or {}
                # 人群/内容限制恒定不出列，界面上没有开关
                path = WT.build(cfg, positions,
                                existing_activity=bool(opts.get("existing_activity")))
            elif cfg.get("mode") == "price_panel":
                # 和资源位投放同一个开关：挂到已有活动就不生成「活动」sheet
                from . import pp_template as PT
                path = PT.build(cfg, existing_activity=bool((options or {}).get("existing_activity")))
            else:
                spec = registry.spec_for(cfg.get("mode"))
                if registry.scopes_for(cfg) and scope != "id_list":
                    return {"ok": False, "error": spec.no_template_hint}
                # 有的 mode 压根不吃 Excel（比如抢会议室，任务清单在界面上填），
                # build_template 是 None —— 直接调会 TypeError，得先兜住
                if spec.build_template is None:
                    return {"ok": False,
                            "error": spec.no_template_hint or "这个配置类型不需要 Excel 模板"}
                path = spec.build_template(form_name)
            usage.record(self.settings, "template_made", form=form_name, entry="webui")
            return {"ok": True, "path": str(path)}
        except Exception as e:
            log.exception("生成模板失败")
            return {"ok": False, "error": str(e)}

    def open_path(self, path: str) -> bool:
        try:
            os.startfile(path)
            return True
        except Exception:
            log.exception("打开失败：%s", path)
            return False

    # ---------------- 准备：载入并检查 ----------------
    def load_and_check(self, form_name: str, data_file: str | None, scope: str | None,
                       options: dict | None = None) -> dict:
        try:
            cfg = self._form_cfg(form_name)
            self.form_name, self.scope = form_name, scope
            s = dict(self.settings)
            s["data_file"] = data_file or ""
            s["resume"] = True     # 「跳过已成功的」由前端勾选框在开跑时决定，这里始终读断点
            s["dmp_scope"] = scope
            s["ab_scope"] = scope
            # 价格策略批量开关：方向 + 选哪些行（keyword/ledger/list）+ 日期区间 + 策略
            s["pt_scope"] = scope
            _o = options or {}
            s["toggle_direction"] = "off" if _o.get("toggle_direction") == "off" else "on"
            s["toggle_params"] = str(_o.get("toggle_params", "") or "")
            s["toggle_strategies"] = str(_o.get("toggle_strategies", "") or "")
            s["toggle_date_from"] = str(_o.get("toggle_date_from", "") or "")
            s["toggle_date_to"] = str(_o.get("toggle_date_to", "") or "")
            # 活动挂哪儿、按哪套策略跑。资源位投放和价格面板配置共用这两个键
            # （界面上就是同一行控件）；别的 mode 用不到。
            if cfg.get("mode") in ("wizard", "price_panel"):
                from . import wizard_strategy as S
                s["wizard_activity"] = dict(options or {}).get("activity") or {}
                s["wizard_strategy"] = S.active_payload(cfg)
            # 原生商广专用：计划名称/转化目标/出价/投放时间/人群，界面上填一次全批共用
            if cfg.get("mode") == "ad_native":
                from . import ad_prep as P
                s["ad_prep"] = P.load(cfg)
            # 价格面板：投放配置页上的「生效渠道」（+ 定向时的 panel_type）。
            # ⚠ 它决定页面是哪一套表单，所以填之前必须拿到。
            if cfg.get("mode") == "price_panel":
                from . import ad_prep as P
                s["pp_prep"] = P.load(cfg)
            # 预定会议室专用：抢占任务清单，界面上填的，不走 data_file
            if cfg.get("mode") == "meeting_reserve":
                from . import meeting_data as MD
                s["meeting_tasks"] = MD.load(cfg)["tasks"]

            self.ui = WebUI(self._window)
            self.runner = registry.spec_for(cfg.get("mode")).make_runner(s, cfg, self.ui)
            rows = self.runner.preview()
            self.preview_rows = rows

            bad = sum(1 for r in rows if r.issues)
            usage.record(self.settings, "load_checked", form=form_name, scope=scope,
                         total=len(rows), bad=bad, entry="webui",
                         # 哪几列没填对 —— 只记列名（模板表头），不记用户填的值。
                         # 这个数直接告诉你该改模板哪一行填写说明
                         bad_fields=(usage.bad_fields(rows) or None))
            return {
                "ok": True,
                "rows": _json_safe([self._row_summary(r) for r in rows]),
                "total": len(rows), "bad": bad,
            }
        except Exception as e:
            log.exception("载入失败")
            return {"ok": False, "error": str(e)}

    @staticmethod
    def _row_summary(r) -> dict:
        # payload 不传回前端：有的（wizard 的 creatives）很大，且前端不需要自己拼 records，
        # 双击详情 / 开跑都是回头问 Python 要，数据不用来回搬两遍。
        return {"index": r.index, "name": r.name, "kind": r.kind,
                "detail_count": r.detail_count, "issues": r.issues, "done": r.done}

    def row_detail(self, index: int) -> dict | None:
        row = next((r for r in self.preview_rows if r.index == index), None)
        if not row:
            return None
        rec = row.payload
        items = rec.get("items") if "items" in rec else rec.get("creatives", [])
        return _json_safe({
            "index": row.index, "name": row.name, "issues": row.issues,
            "header": rec.get("header", {}), "items": items or [],
        })

    def pt_ledger_view(self, form_name: str) -> dict:
        """「价格策略批量开关」卡里那份台账：本工具「价格策略配置」配过哪些。

        strategies：出现过的策略（新→旧），给「策略」下拉当选项。
        recent：最近几批，给人看一眼「都记了些啥」。
        """
        try:
            cfg = self._form_cfg(form_name)
            name = cfg.get("ledger")
            if not name:
                return {"ok": True, "strategies": [], "recent": [], "path": ""}
            from . import pt_ledger as PL
            recent = [{
                "at": b.get("at", ""),
                "strategy": b.get("strategy_name") or b.get("strategy_id", ""),
                "strategy_id": b.get("strategy_id", ""),
                "count": len(b.get("names") or []),
                "names": (b.get("names") or [])[:8],
            } for b in PL.load(name)[:12]]
            return _json_safe({
                "ok": True,
                "strategies": PL.strategies(name),
                "recent": recent,
                "path": PL.path(name),
            })
        except Exception as e:
            log.exception("读台账失败")
            return {"ok": False, "error": str(e)}

    def clear_state(self, form_name: str) -> dict:
        try:
            cfg = self._form_cfg(form_name)
            registry.spec_for(cfg.get("mode")).make_runner(
                dict(self.settings), cfg, WebUI(self._window)).clear_state()
            return {"ok": True}
        except Exception as e:
            log.exception("清除断点失败")
            return {"ok": False, "error": str(e)}

    def submit_feedback(self, payload: dict | None) -> dict:
        """用户反馈：报告问题（可附最近日志）/ 功能建议。发到企微群。

        ⚠ 内容是用户自己写、自己点发的，不是埋点 —— 但日志段可能带业务字样，
          所以界面上那个「附上运行日志」的勾选默认给用户看得见、可关掉。
        """
        try:
            from . import report, usage
            p = payload or {}
            kind = "报告问题" if str(p.get("kind")) == "issue" else "功能建议"
            text = str(p.get("text", "")).strip()
            if not text:
                return {"ok": False, "error": "反馈内容是空的"}
            tail = str(p.get("log", "")).strip()

            lines = [f"【配置助手 · {kind}】",
                     f"版本 {usage._app_version()}　指纹 {usage._uid()}"]
            if self.form_name:
                lines.append(f"配置类型 {self.form_name}")
            lines += ["——", text[:800]]
            if tail:
                lines += ["——", "最近日志：", tail[:1200]]

            ok = report.send_feedback(self.settings, "\n".join(lines))
            return {"ok": True} if ok else {
                "ok": False, "error": "没发出去 —— 检查网络，或稍后再试"}
        except Exception as e:
            log.exception("提交反馈失败")
            return {"ok": False, "error": str(e)}

    def open_output_dir(self) -> bool:
        p = ROOT / "output"
        p.mkdir(parents=True, exist_ok=True)
        subprocess.Popen(["explorer", str(p)])
        return True

    # ---------------- 自制配置类型（mode: flow）----------------
    def flow_list(self) -> list:
        """《自制配置类型》分组里那些，加上状态。"""
        from . import flow_data as FD
        out = []
        for name, cfg in FD.list_all():
            f = cfg["_flow"]
            out.append(_json_safe({
                "name": name, "status": f.get("status", "draft"),
                "steps": len(f.get("steps") or []), "created_at": f.get("created_at", ""),
                "source_url": f.get("source_url", ""), "loop": FD.has_loop(f),
                "columns": FD.columns(f),
            }))
        return out

    def flow_get(self, name: str) -> dict:
        from . import flow_data as FD
        try:
            f = FD.load(name)
            return _json_safe({"ok": True, "flow": f, "issues": FD.validate(f),
                               "columns": FD.columns(f)})
        except Exception as e:
            return {"ok": False, "error": str(e)}

    def flow_new(self, name: str, url: str) -> dict:
        from . import flow_data as FD
        name = str(name or "").strip()
        if not name:
            return {"ok": False, "error": "先起个名字"}
        if FD.exists(name) or formcfg.path_for(name).exists():
            return {"ok": False, "error": f"「{name}」这个名字已经有了，换一个"}
        try:
            from . import usage
            FD.save({"name": name, "source_url": str(url or "").strip(), "status": "draft",
                     "created_by": usage._uid(),
                     "data": {"source": "none", "columns": []}, "steps": []})
            return {"ok": True}
        except Exception as e:
            log.exception("新建自制工作流失败")
            return {"ok": False, "error": str(e)}

    def flow_save(self, name: str, doc: dict) -> dict:
        from . import flow_data as FD
        try:
            d = dict(doc or {})
            d["name"] = name
            path = FD.save(d)
            f = FD.load(name)
            return _json_safe({"ok": True, "path": path, "issues": FD.validate(f),
                               "columns": FD.columns(f)})
        except Exception as e:
            log.exception("存自制工作流失败")
            return {"ok": False, "error": str(e)}

    def flow_delete(self, name: str) -> dict:
        from . import flow_data as FD
        try:
            p = FD.path_for(name)
            if p.exists():
                p.unlink()
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    # ---- 录制 ----
    def flow_start_record(self, name: str) -> dict:
        from . import flow_data as FD
        from .browser import Browser
        from .flow_record import FlowRecorder
        if self._flow_rec:
            return {"ok": False, "error": "已经在录了，先点「完成」或「停止录制」"}
        if not chrome.is_connected(self.settings["cdp_url"]):
            return {"ok": False, "error": "浏览器没连上，先启动浏览器并登录"}
        try:
            f = FD.load(name)
            b = Browser(self.settings["cdp_url"], self.settings["timeout"])
            b.__enter__()
            if f.get("source_url"):
                try:
                    b.page.goto(f["source_url"], wait_until="domcontentloaded")
                except Exception:
                    pass
            b.front()
            rec = FlowRecorder(b.page, self.settings["timeout"])
            rec.start()
            self._flow_browser, self._flow_rec = b, rec
            return {"ok": True}
        except Exception as e:
            log.exception("开始录制失败")
            self._flow_cleanup()
            return {"ok": False, "error": str(e)}

    def flow_record_status(self) -> dict:
        rec = self._flow_rec
        if not rec:
            return {"running": False, "done": False, "steps": 0}
        try:
            return {"running": True, "done": bool(rec.done), "steps": len(rec.steps)}
        except Exception:
            return {"running": True, "done": False, "steps": 0}

    def flow_stop_record(self, name: str) -> dict:
        from . import flow_data as FD
        rec = self._flow_rec
        if not rec:
            return {"ok": False, "error": "没有在录"}
        try:
            steps = rec.stop()
        except Exception as e:
            log.exception("停止录制出错")
            steps = getattr(rec, "steps", [])
        finally:
            self._flow_cleanup()
        try:
            f = FD.load(name)
            f["steps"] = steps
            if not f.get("source_url") and steps and steps[0].get("op") == "goto":
                f["source_url"] = steps[0]["url"]
            FD.save(f)
            return _json_safe({"ok": True, "flow": FD.load(name),
                               "issues": FD.validate(FD.load(name))})
        except Exception as e:
            log.exception("录制结果存盘失败")
            return {"ok": False, "error": str(e)}

    def _flow_cleanup(self):
        b = self._flow_browser
        self._flow_browser = self._flow_rec = None
        if b:
            try:
                b.__exit__(None, None, None)
            except Exception:
                pass

    # ---- 送审 ----
    def flow_submit(self, name: str) -> dict:
        from . import flow_data as FD, flow_review
        try:
            f = FD.load(name)
            issues = FD.validate(f)
            if issues:
                return {"ok": False, "error": "还有问题没解决：" + "；".join(issues[:3])}
            csv_text = ""
            try:
                csv_text = Path(self.settings["result_file"]).read_text(encoding="utf-8-sig")
            except OSError:
                pass
            res = flow_review.submit(self.settings, f, csv_text)
            if res.get("ok"):
                f["status"] = "submitted"
                FD.save(f)
            return _json_safe(res)
        except Exception as e:
            log.exception("送审失败")
            return {"ok": False, "error": str(e)}

    def flow_mark_tested(self, name: str) -> dict:
        from . import flow_data as FD
        try:
            f = FD.load(name)
            if f.get("status") == "draft":
                f["status"] = "tested"
                FD.save(f)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "error": str(e)}

    @staticmethod
    def _flow_template(name: str, columns: list) -> str:
        from openpyxl import Workbook
        from .paths import user_path
        wb = Workbook()
        ws = wb.active
        ws.append(list(columns))
        for c in range(1, len(columns) + 1):
            ws.cell(row=1, column=c).font = ws.cell(row=1, column=c).font.copy(bold=True)
        p = user_path("data", f"{name}_数据.xlsx")
        p.parent.mkdir(parents=True, exist_ok=True)
        wb.save(p)
        return str(p)

    # ---------------- 执行 ----------------
    def start_run(self, mode: str, skip_done: bool) -> dict:
        if not self.runner or not self.preview_rows:
            return {"ok": False, "error": "还没有载入数据，请先在「准备」页载入并检查"}
        if not chrome.is_connected(self.settings["cdp_url"]):
            return {"ok": False, "error": "浏览器没连上，请先启动浏览器并登录"}

        good = [r for r in self.preview_rows if not r.issues]
        kept = [r for r in good if not (skip_done and r.done)]
        if not kept:
            return {"ok": False, "error": "所有数据要么有问题、要么已完成"}
        return self._run(kept, mode)

    def retry_rows(self, indices: list, mode: str) -> dict:
        """只重跑「失败清单」里选中的那几条，复用同一个 runner 实例。"""
        if not self.runner:
            return {"ok": False, "error": "还没有载入数据，请先在「准备」页载入并检查"}
        if not chrome.is_connected(self.settings["cdp_url"]):
            return {"ok": False, "error": "浏览器没连上，请先启动浏览器并登录"}
        wanted = set(indices or [])
        rows = [r for r in self.preview_rows if r.index in wanted]
        if not rows:
            return {"ok": False, "error": "没找到要重跑的行"}
        # retry_of 指回上一次运行：重跑的量不能再算一遍战绩（同一批活干了两遍）
        return self._run(rows, mode, retry_of=self.last_run_id)

    def _run(self, kept, mode: str, retry_of: str | None = None) -> dict:
        records = [r.payload for r in kept]
        self.runner.s["dry_run"] = (mode == "dry")
        self.runner.auto = (mode == "auto")
        self.ui = WebUI(self._window)
        self.runner.ui = self.ui
        self._run_kept_rows = kept

        run_id = usage.new_run_id()
        self.last_run_id = run_id

        def work():
            t0 = time.monotonic()
            try:
                self.last_results = self.runner.run(records)
            except Exception as e:
                log.exception("运行出错")
                self.ui.log(f"运行中断：{e}", "error")
                self.last_results = []
            finally:
                self._record_run(run_id, mode, retry_of, len(records),
                                 time.monotonic() - t0, self.ui.wait_seconds)
                self._sync_sheet_async()
                self.ui._push("onRunDone", self._failed_summary())

        self.worker = threading.Thread(target=work, daemon=True)
        self.worker.start()
        return {"ok": True, "total": len(records)}

    def _record_run(self, run_id, mode, retry_of, total, seconds, wait_seconds):
        """一次运行落一条埋点。⚠ 只记数量和状态，业务内容一个字都不带 ——
        这份文件是要汇总给所有人看的，理由见 src/usage.py 开头。"""
        counts = usage.count_status(self.last_results)
        usage.record(
            self.settings, "run_finished",
            run_id=run_id, retry_of=retry_of,
            form=self.form_name, mode=mode, scope=self.scope,
            total=total, seconds=round(seconds, 1),
            # 「等人点确认」的时间单独记：逐条确认模式下这部分不能算机器代劳
            wait_seconds=round(wait_seconds, 1),
            # 是不是中途点了停止。以前只能靠 ok+failed+skipped < total 反推，太绕
            stopped=(True if (self.ui and self.ui.stop_flag.is_set()) else None),
            chrome=usage.chrome_version(self.settings.get("cdp_url")),
            entry="webui", **counts,
            # 失败分成哪几类、卡在哪一层（都是固定枚举，不含错误原文）
            **usage.fail_detail(self.last_results),
            # 单条耗时分位数：区分「整体慢」和「个别卡死」
            **usage.percentiles(self.ui.item_seconds if self.ui else []),
        )

    def _sync_sheet_async(self, quiet: bool = False):
        """把自己欠着的那几周发到统计群。跑完一轮之后、启动时各调一次。

        ⚠ 2026-08-21 换了通道：原来是开浏览器操作企微文档（十几秒、要登录态、
          页面结构一变就废），现在是一个 urllib POST，几百毫秒，前端完全无感知。
          原委见 src/report.py 的文件头。
        ⚠ 仍然放后台线程：内网偶尔抽风，三秒超时也不该挡着界面上「跑完了」的提示。
        ⚠ 不吞掉失败的**事实**：异常不往上抛（统计不能挡业务），但要记进
          self._sync_error，首页会显示「还有 N 周没上报」。欠着的周不会丢 ——
          usage.report_rows 按差异补报，下次连上就全补齐。
        """
        from . import report
        if not report.enabled(self.settings):
            return
        if self._syncing:
            return

        def work():
            self._syncing = True
            try:
                names = [f["name"] for f in self.list_forms()]
                res = report.push(self.settings, names,
                                  (self.settings.get("usage") or {}).get("nickname") or "")
                self._sync_error = res.get("error") or ""
                if res.get("sent"):
                    log.info("统计已上报 %d 周", res["sent"])
                if res.get("failed") and not quiet:
                    self._toast(f"统计没能上报：{self._sync_error}"
                                f"　（数据没丢，下次会自动补）", "warn")
            except Exception as e:
                self._sync_error = str(e)
                log.warning("统计上报失败（不影响本次运行，下次会补报）", exc_info=True)
            finally:
                self._syncing = False

        threading.Thread(target=work, daemon=True).start()

    def catch_up_report(self):
        """启动时补一次上报。

        ⚠ 为什么启动时也要来一次，而不是只在跑完之后：跑完那次是后台线程，
          **关窗口就跟着没了** —— 人点完「跑完了」就叉掉，上报根本走不完。
          实测「有人用了一天、表里没数」多半是这个。启动补这一趟，
          昨天欠的今天开机就补上了。
        ⚠ 不再判断浏览器开没开 —— 新通道是纯 HTTP，和 Chrome 一点关系都没有。
        """
        try:
            from . import report
            if not report.enabled(self.settings):
                return
            names = [f["name"] for f in self.list_forms()]
            if not usage.pending_weeks(self.settings, names):
                return
            self._sync_sheet_async(quiet=True)
        except Exception:
            log.warning("启动补报失败", exc_info=True)

    def _failed_summary(self) -> dict:
        """把刚跑完的失败项整理成 [{index, name, error}]，重跑单条时按 index 找回原始行。

        ⚠ 这里假设 self.last_results 和 self._run_kept_rows 位置一一对应——
          「跳过已成功的」勾选时（默认勾选）这个假设总成立：records 里不会有
          runner 自己状态也认为"已完成"的行，run() 内部就不会走"已完成过，
          跳过"这条不 append 结果的分支，位置不会错位。不勾选「跳过已成功的」
          时，如果这次要跑的行里混了真的已经完成过的，位置可能对不上——这时
          宁可不做逐行关联（misaligned=True，前端只展示"跑完了"，不给重跑
          按钮），也不能把错误安在错的行上。
        """
        kept, results = self._run_kept_rows, self.last_results
        if len(results) != len(kept):
            return {"failed": [], "misaligned": bool(kept)}
        failed = [{"index": row.index, "name": row.name, "error": res.get("错误", "")}
                  for row, res in zip(kept, results) if res.get("状态") == "failed"]
        return {"failed": failed, "misaligned": False}

    def pause_run(self) -> dict:
        if not self.ui:
            return {"paused": False}
        if self.ui.pause_flag.is_set():
            self.ui.pause_flag.clear()
            return {"paused": False}
        self.ui.pause_flag.set()
        return {"paused": True}

    def stop_run(self) -> bool:
        if self.ui:
            self.ui.stop_flag.set()
            self.ui.pause_flag.clear()
        return True

    def answer(self, value):
        """核对确认 / 失败继续 两种弹窗共用的回传口。"""
        if not self.ui:
            return
        try:
            self.ui.answer.put_nowait(value)
        except queue.Full:
            pass


def main():
    try:
        settings = _load_settings()
    except Exception as e:
        # 没有控制台时静默失败没法排查，至少弹一个原生消息框
        notify.popup("启动失败", f"{e}\n\n程序目录：{ROOT}", warn=True)
        return

    # 「打开了但一条没跑」= 有人试了没跑起来。这个差值不埋就永远看不见
    usage.record(settings, "app_open", entry="webui", dpi=usage.dpi_scale())

    api = Api(settings)
    index = resource("assets", "webui", "index.html")
    if not index:
        notify.popup("启动失败", f"找不到界面文件：assets/webui/index.html\n程序目录：{ROOT}", warn=True)
        return

    window = webview.create_window(
        "大会员业务后台 配置助手",
        url=str(index),
        js_api=api,
        width=1020, height=820, min_size=(940, 720),
    )
    api._window = window
    # 启动补报：昨天没报上去的，今天开机自动补（纯 HTTP，和浏览器无关）
    threading.Timer(3.0, api.catch_up_report).start()
    webview.start()


if __name__ == "__main__":
    main()
