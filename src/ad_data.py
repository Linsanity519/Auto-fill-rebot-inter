"""原生商广的数据读取：把一张平表读成「单元 → 创意」两层。

⚠ 只服务 mode: ad_native。资源位投放走 wizard_data，老配置走 datasource，三边互不影响。

模板是一张平表，一行 = 一条创意：

    内容            avid              素材标题        素材描述    封面
    牧神记年番      116453323900575   xxxx            牧神记年番
    牧神记年番      114096477310907   yyyy            牧神记年番
    影后            114096477310907   zzzz            影后

「内容」相同的行聚成一个单元（同一个内容为一个聚类），单元名称按 yaml 的
name_template 拼。页面上一个单元最多挂 10 条创意（抽屉里写死「已选 n/10」），
超了就自动拆成多个单元，名字用 overflow_template 加 _0、_1、_2… 顺延区分，不然会重名。

⚠ 同一个 avid 写多行是合法的，而且不能合并 —— 封面和素材标题要求一一对应，
  同一条视频配不同封面 + 不同标题时就是两条独立的创意。
  实测过：「添加稿件/视频」抽屉里把同一个 avid 分两次加进来，页面确实会生成
  两个独立创意块（data-id 分别是 c_0_<avid>_… 和 c_1_<avid>_…）。
  所以这里记一个 _seq（这个 avid 在本单元里的第几次出现），填写时靠它认块。

⚠ 聚类不要求同一个内容的行必须挨着 —— 运营手工整理的表经常是乱序的，
  这里按首次出现顺序归组，中间插了别的内容也能正确合并。

⚠ 隐藏的行和列直接忽略（见 _rows）。

封面列两种填法都支持（和资源位投放一致）：
  · 单元格里写本地路径    → 直接用
  · 图片直接贴在单元格上  → 从 xlsx 里抽出来存到 output/_images/ 再用
"""
from __future__ import annotations

import logging
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

# ⚠ 抽图那两个函数和资源位投放一模一样（openpyxl 的 _images 锚点解析很啰嗦），
#   直接复用 wizard_data 里的实现，不重写第二份。
from .wizard_data import DataError, _apply_images, _extract_images

log = logging.getLogger(__name__)

AVID = "avid"


def _grouping(cfg: dict) -> dict:
    return cfg.get("grouping") or {}


def key_column(cfg: dict) -> str:
    return _grouping(cfg).get("key_column", "内容")


def max_creatives(cfg: dict) -> int:
    return int(_grouping(cfg).get("max_creatives", 10))


def columns(cfg: dict) -> list[dict]:
    return [dict(c) for c in (cfg.get("columns") or [])]


def unit_name(cfg: dict, key: str, seq: int, today: str, split: bool = False) -> str:
    """单元名称。

    seq 从 0 起。split=True（同一个内容拆出了多个单元）时才带序号后缀，
    并且从 0 开始顺延：…_0、…_1、…_2。只出一个单元时不加后缀。
    """
    g = _grouping(cfg)
    tpl = g.get("name_template", "【{内容}】_{日期}")
    if split:
        tpl = g.get("overflow_template") or (tpl + "_{序号}")
    return tpl.format(**{"内容": key, "日期": today, "序号": seq})


# ---------------------------------------------------------------- 读表
def _hidden_cols(ws) -> set[int]:
    """被隐藏的列号集合（1 起）。

    ⚠ openpyxl 的 column_dimensions 一条记录可能横跨好几列（min/max），
      不能只按 key 那一个字母算，否则成组隐藏的列会漏掉。
    """
    out: set[int] = set()
    for key, dim in ws.column_dimensions.items():
        if not dim.hidden:
            continue
        try:
            lo = dim.min or column_index_from_string(key)
            hi = dim.max or lo
        except ValueError:
            continue
        out.update(range(int(lo), int(hi) + 1))
    return out


def _hidden_rows(ws) -> set[int]:
    out: set[int] = set()
    for r, dim in ws.row_dimensions.items():
        if dim.hidden:
            out.add(int(r))
    return out


def _rows(path: str) -> tuple[list[dict], dict]:
    """读第一个 sheet 成 [{列名: 值}]，全空行跳过。返回 (行, {列名: 列号})。

    ⚠ 隐藏的行和列一律当不存在：运营的素材表里常有「已下线」「备用」之类
      藏起来的行，还有算 CTR 用的中间列 —— 眼睛看不见的东西不该被投出去。
    """
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        raise DataError(f"打不开数据文件：{e}") from e

    ws = wb[wb.sheetnames[0]]
    skip_cols = _hidden_cols(ws)
    skip_rows = _hidden_rows(ws)
    if skip_cols or skip_rows:
        log.info("忽略隐藏内容：%d 列、%d 行", len(skip_cols), len(skip_rows))

    headers, idx = [], {}
    for i in range(1, ws.max_column + 1):
        if i in skip_cols:
            headers.append("")          # 占位，保持列号对齐
            continue
        v = ws.cell(1, i).value
        h = str(v).strip() if v is not None else ""
        headers.append(h)
        if h:
            idx[h] = i

    out = []
    for r in range(2, ws.max_row + 1):
        if r in skip_rows:
            continue
        rec, empty = {}, True
        for i, h in enumerate(headers, 1):
            if not h:
                continue
            v = ws.cell(r, i).value
            # ⚠ avid 是 15 位整数，openpyxl 读成 int 后 str() 不会变科学计数法，
            #   但用户从别处粘过来可能是 float（1.1645332390058e+14），这里统一收干净
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            s = "" if v is None else str(v).strip()
            rec[h] = s
            if s:
                empty = False
        if not empty:
            rec["_row"] = r
            out.append(rec)
    return out, idx


def load(data_file: str, cfg: dict, settings: dict | None = None) -> dict:
    """读成 {"units": [...]}。每个单元 {key, name, row, creatives:[...]}。"""
    if not data_file:
        raise DataError("还没选数据文件")

    log.info("读数据文件：%s", data_file)
    rows, idx = _rows(data_file)
    if not rows:
        # ⚠ 报错必须带上文件名。分发包的 data/ 里躺着一份「生成出来的空模板」，
        #   选文件的对话框默认又是从那儿打开的 —— 最容易犯的错就是选中空模板，
        #   而原来的报错只有「一行数据都没有」，看不出选错了文件。
        name = Path(data_file).name
        hint = ("　这份是生成出来的空模板（只有表头），"
                "请选你自己填过数据的那份表。" if "模板" in name else
                "　第 1 行是表头，数据从第 2 行开始；整行被隐藏的也会被跳过。")
        raise DataError(f"「{name}」里一行数据都没有。\n{hint}\n完整路径：{data_file}")

    kcol = key_column(cfg)
    known = {c["name"] for c in columns(cfg)}
    missing = [c["name"] for c in columns(cfg)
               if c.get("required") and c["name"] not in idx]
    if missing:
        raise DataError(f"「{Path(data_file).name}」表头缺少必填列：{'、'.join(missing)}。\n"
                        f"　现有列：{'、'.join(k for k in idx) or '（空）'}"
                        f"　—— 整列被隐藏的列也算不存在。")

    # 贴在单元格里的封面图 → 落地成文件路径
    imgs = _extract_images(data_file)
    sheet = next(iter(imgs), "")
    if imgs:
        _apply_images(rows, idx, imgs.get(sheet, {}), sheet)

    _apply_fallbacks(rows, columns(cfg))

    today = datetime.now().strftime("%Y%m%d")
    cap = max_creatives(cfg)

    # 按首次出现顺序归组
    order: list[str] = []
    buckets: dict[str, list[dict]] = {}
    for rec in rows:
        key = rec.get(kcol, "").strip()
        if not key:
            raise DataError(f"第{rec['_row']}行的「{kcol}」是空的，没法归到哪个单元")
        if key not in buckets:
            buckets[key] = []
            order.append(key)
        buckets[key].append({k: v for k, v in rec.items() if k in known or k == "_row"})

    units = []
    for key in order:
        items = _number_repeats(buckets[key])
        # 超过 10 条创意就切片，每片一个单元
        split = len(items) > cap
        for n, start in enumerate(range(0, len(items), cap)):
            chunk = items[start:start + cap]
            units.append({
                "key": key,
                "name": unit_name(cfg, key, n, today, split),
                "row": chunk[0]["_row"],
                "creatives": chunk,
            })
    return {"units": units}


def _apply_fallbacks(rows: list[dict], cols: list[dict]):
    """列上写了 fallback: 某列 的，留空时就抄那一列的值。

    「素材描述」留空抄「内容」—— 两列绝大多数时候都是同一个剧集名，
    让人填两遍纯属重复劳动。只有剧集名超过页面 10 字上限时才需要单独填个短的。
    """
    pairs = [(c["name"], c["fallback"]) for c in cols if c.get("fallback")]
    if not pairs:
        return
    for rec in rows:
        for name, src in pairs:
            if not str(rec.get(name, "")).strip():
                rec[name] = rec.get(src, "")


def _number_repeats(items: list[dict]) -> list[dict]:
    """给每一行标上 _seq：这个 avid 在本单元里是第几次出现（从 0 起）。

    ⚠ 页面上同一个 avid 加两次会得到两个创意块，先加的排在前面。
      填写时按「匹配这个 avid 的第 _seq 个块」定位，不依赖块与块之间的整体顺序。
    """
    seen: dict[str, int] = {}
    out = []
    for rec in items:
        avid = str(rec.get(AVID, "")).strip()
        rec = dict(rec)
        rec["_seq"] = seen.get(avid, 0)
        seen[avid] = rec["_seq"] + 1
        out.append(rec)
    return out


def add_passes(creatives: list[dict]) -> list[list[dict]]:
    """把一个单元的创意分成几趟加：第 n 趟收各 avid 的第 n 次出现。

    ⚠ 抽屉一次「确定」里同一个 avid 只能勾一次（勾了就是选中态），
      所以重复的 avid 必须分多趟。按出现次序分趟，既最少开抽屉次数，
      又保证「先加的是 _seq 小的」这个前提成立。
    """
    passes: list[list[dict]] = []
    for c in creatives:
        i = int(c.get("_seq", 0))
        while len(passes) <= i:
            passes.append([])
        passes[i].append(c)
    return passes


def validate(cfg: dict, data: dict) -> list[str]:
    """跑之前的体检。返回人话的问题清单，每条都带行号。"""
    issues = []
    seen_names: dict[str, int] = {}
    for u in data["units"]:
        head = f"「{u['name']}」"
        if u["name"] in seen_names:
            issues.append(f"{head} 单元名和第{seen_names[u['name']]}行重名，"
                          f"后台会建出两个同名单元")
        seen_names[u["name"]] = u["row"]

        for c in u["creatives"]:
            r = c["_row"]
            avid = str(c.get(AVID, "")).strip()
            if not avid:
                issues.append(f"{head} 第{r}行没填 avid")
            elif not re.fullmatch(r"\d+", avid):
                issues.append(f"{head} 第{r}行的 avid「{avid}」不是纯数字")

            for f in (cfg.get("creative") or {}).get("fields", []):
                name = f["name"]
                val = str(c.get(name, "")).strip()
                if not val:
                    if f.get("required"):
                        issues.append(f"{head} 第{r}行「{name}」没填")
                    continue
                mx = f.get("max")
                if mx and len(val) > int(mx):
                    # 把原文带上：这条基本都是「剧集名比 10 个字长」，
                    # 带上原文才能直接照着改，不用再回表里翻行号
                    issues.append(f"{head} 第{r}行「{name}」{len(val)} 字，"
                                  f"超过页面上限 {mx} 字：{val}")
    return issues
