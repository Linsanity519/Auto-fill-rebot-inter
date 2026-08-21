"""读 Excel / CSV，并按「分组」列把多行合并成「一条记录 + 多个明细项」。

长表格式（推荐）——一条配置有几项就写几行，用「分组」列归并：

  分组 | 价格人群名称 | 人群选组 | 平台          | 优先级 | 限制类型 | 会员卡种 | 价格均值 | 价格区间下限 | 价格区间上限
  1    | 新客低价     | 不限     | iPhone,Android| 10     | 常规均价 | 连续包年 | 88       | 80           | 100
  1    |              |          |               |        |          | 连续包月 | 15       | 12           | 20
  2    | 老客召回     | 不限     | pc            | 20     | 常规均价 | 年度大会员| 148     | 140          | 160

同一分组里，主表字段只需在第一行填，后续行留空即可。
没有「分组」列时，退化成一行一条记录、每条一个明细项。
"""
import csv
import datetime as _dt
from pathlib import Path

from openpyxl import load_workbook

# ⚠ 这里以前用 pandas.read_excel/read_csv(dtype=str)。pandas + numpy 在打包后要占
#   90MB，而全项目只用到下面这一个函数 —— 为了让"自动更新"每次只下几百 KB 而不是
#   上百 MB，改用已经是依赖的 openpyxl + 标准库 csv。行为对齐原来的 dtype=str：
#   一律给字符串、空单元格给 ""、列名 strip。


def _cell_text(v) -> str:
    """单元格 → 字符串。对齐 pandas dtype=str 的口径，另外修掉两个它的糟糕输出。"""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, bool):
        return "True" if v else "False"
    # ⚠ Excel 里所有数字都是 double。pandas dtype=str 会把整数 88 渲染成 "88.0"，
    #   填进"价格""优先级"这种框里是错的，所以整数值一律去掉小数尾巴。
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    if isinstance(v, int):
        return str(v)
    # ⚠ 日期同理：pandas 给的是 "2026-08-21 00:00:00"，日期控件不认。
    #   只有确实带时分秒时才带上。
    if isinstance(v, _dt.datetime):
        return v.strftime("%Y-%m-%d" if (v.hour, v.minute, v.second) == (0, 0, 0)
                          else "%Y-%m-%d %H:%M:%S")
    if isinstance(v, _dt.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, _dt.time):
        return v.strftime("%H:%M:%S")
    return str(v)


def _rows_to_records(rows) -> list[dict]:
    """第一行当表头，其余按表头组装成 dict。空表头的列丢掉，全空的行丢掉。"""
    rows = iter(rows)
    try:
        header_row = next(rows)
    except StopIteration:
        return []

    # 列名 strip；表头为空的列不要（pandas 会叫它 Unnamed: N，下游也用不上）
    cols = [(i, _cell_text(c).strip()) for i, c in enumerate(header_row)]
    cols = [(i, name) for i, name in cols if name]
    if not cols:
        return []

    records = []
    for row in rows:
        rec = {name: _cell_text(row[i]) if i < len(row) else "" for i, name in cols}
        # 尾部空行：openpyxl 常会多吐一堆，pandas 是不给的
        if any(v != "" for v in rec.values()):
            records.append(rec)
    return records


def load_table(path: str, sheet_name=None) -> list[dict]:
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"数据文件不存在：{p.resolve()}")

    suffix = p.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        # read_only=True：大表不必整本读进内存；data_only=True：公式取算好的值，
        # 不然拿到的是 "=A1*2" 这种公式串。
        wb = load_workbook(p, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
            return _rows_to_records(ws.iter_rows(values_only=True))
        finally:
            wb.close()
    elif suffix == ".csv":
        # utf-8-sig：Excel 存出来的 CSV 带 BOM，不吃掉的话第一个列名会多个 ﻿
        with p.open("r", encoding="utf-8-sig", newline="") as f:
            return _rows_to_records(csv.reader(f))
    else:
        raise ValueError(f"不支持的文件类型：{p.suffix}")


def item_field_names(list_cfg: dict) -> list[str]:
    """明细字段名；有 variants 时取所有变体的并集（保持出现顺序）。"""
    if not list_cfg:
        return []
    groups = list(list_cfg["variants"].values()) if "variants" in list_cfg else [list_cfg["fields"]]
    names = []
    for g in groups:
        for f in g:
            if f["name"] not in names:
                names.append(f["name"])
    return names


def header_field_names(form_cfg: dict) -> list[str]:
    """主表字段名，含 reveals 里的条件字段。"""
    names = []
    for f in form_cfg["fields"]:
        names.append(f["name"])
        for subs in (f.get("reveals") or {}).values():
            for s in subs:
                if s["name"] not in names:
                    names.append(s["name"])
    return names


def build_records(rows: list[dict], form_cfg: dict, group_key: str = "分组") -> list[dict]:
    """把扁平的表格行组装成 [{header: {...}, items: [{...}], source_rows: [i,...]}]"""
    list_cfg = form_cfg.get("list")
    header_names = header_field_names(form_cfg)
    item_names = item_field_names(list_cfg)

    # 没配重复行，或表里没有分组列 → 一行一条
    if not list_cfg or not rows or group_key not in rows[0]:
        return [
            {"header": r, "items": [r] if item_names else [], "source_rows": [i]}
            for i, r in enumerate(rows)
        ]

    records, index = [], {}
    for i, r in enumerate(rows):
        gid = str(r.get(group_key, "")).strip() or f"__row{i}"
        if gid not in index:
            index[gid] = {
                "header": {k: r.get(k, "") for k in header_names},
                "items": [],
                "source_rows": [],
            }
            records.append(index[gid])
        rec = index[gid]
        # 后续行里主表字段若非空则覆盖（允许中途修正）
        for k in header_names:
            if str(r.get(k, "")).strip():
                rec["header"][k] = r[k]
        if any(str(r.get(k, "")).strip() for k in item_names):
            rec["items"].append({k: r.get(k, "") for k in item_names})
        rec["source_rows"].append(i)

    return records


def check_columns(rows: list[dict], form_cfg: dict) -> list[str]:
    """返回缺失的必填列名。

    只校验无条件必填的主表字段。条件字段（reveals）和明细字段随变体变化，
    在实际填写时按行校验——这里一刀切会误报。
    """
    if not rows:
        return []
    cols = set(rows[0].keys())
    need = [f["name"] for f in form_cfg["fields"] if f.get("required")]
    return [n for n in need if n not in cols]
