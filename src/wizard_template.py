"""wizard 模式的 Excel 模板生成：按选好的资源位动态出列。

⚠ 只服务 mode: wizard。老配置走 src/template.py，两边互不影响。

产出结构：
  Sheet「活动」      —— 一行，活动层字段；挂到已有活动时不生成这个 sheet
  Sheet「资源位_xxx」—— 每个资源位一个：单元层 + 创意层合在同一张表里
  Sheet「填写说明」  —— 每列什么意思、策略中心当前配了什么

一行 = 一个单元 + 它的一条创意。
同一个单元要挂第 2、3 条创意：可以继续填写相同的「单元名称」，
也可以紧接着另起一行并留空「单元名称」，只填创意列。

⚠ 单元层里被「策略中心」接管的字段（生效平台/投放流量池/创意赛马……）不出列，
  执行时由 wizard_strategy 统一补。要改这个范围，动 yaml 的 strategy_groups。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import wizard_schema as W
from . import wizard_strategy as S
from . import xlsx_kit as X

# 颜色统一在 src/xlsx_kit.py 里定义
REQ_FILL = X.FILLS["req"]     # 必填 浅黄
OPT_FILL = X.FILLS["opt"]     # 选填 浅灰
KEY_FILL = X.FILLS["key"]     # 关联键 浅蓝
COND_FILL = X.FILLS["cond"]   # 条件列 浅橙

UNIT_NAME = "单元名称"
CREATIVE_PREFIX = "创意·"

# sheet 名上限 31 字符，且不能含 : \ / ? * [ ]
BAD = r':\/?*[]'


def _safe(name: str, prefix: str) -> str:
    clean = "".join(ch for ch in name if ch not in BAD)
    return f"{prefix}{clean}"[:31]


def sheet_name(position: str) -> str:
    return _safe(position, "资源位_")


def columns_for_sheet(cfg: dict, position: str) -> list[dict]:
    """一个资源位 sheet 的列清单。

    每项 {title, field, role}：title 是表头文字（创意列带前缀），
    field 是原始字段定义（说明/下拉/校验都从它来），role 决定表头配色。
    """
    cols = [{"title": UNIT_NAME, "field": None, "role": "key"}]
    for f in W.unit_columns_for_template(cfg, position):
        if f["name"] == UNIT_NAME:
            continue                     # 已经作为关联键放在首列了
        cols.append({"title": f["name"], "field": f, "role": "unit"})
    for f in W.columns_for(cfg, position, W.STEP_CREATIVE):
        cols.append({"title": CREATIVE_PREFIX + f["name"], "field": f, "role": "creative"})
    return cols


def _header(ws, cols: list[dict]):
    for i, c in enumerate(cols, 1):
        title, f = c["title"], c["field"]
        cell = ws.cell(row=1, column=i, value=title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if c["role"] == "key":
            cell.fill = KEY_FILL
        elif f and f.get("_when"):
            cell.fill = COND_FILL
        elif f and str(f.get("default", "")).strip():
            cell.fill = OPT_FILL          # 有固定值的列不用填，别标成必填的黄
        elif f and f.get("required"):
            cell.fill = REQ_FILL
        else:
            cell.fill = OPT_FILL
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(30, len(title) * 2.2))

        if f:
            note = W.describe(f)
            if note:
                cell.comment = Comment(note, "配置助手")
            # ⚠ 日期列先把单元格格式设成时间：用户敲「2026-09-24 10:00」时 Excel 会
            #   存成真正的时间值，读出来就是标准写法。不设的话存成文本，
            #   「9.24 10:00:00」这种也能填进去，跑到页面上才发现日期控件不认。
            if f.get("type") in W.DATE_TYPES:
                for r in range(2, 501):
                    ws.cell(row=r, column=i).number_format = "yyyy-mm-dd hh:mm:ss"
        elif title == UNIT_NAME:
            cell.comment = Comment(
                "一行 = 一个单元 + 一条创意。\n"
                "同一个单元要多条创意：可重复填写相同名称，\n"
                "或下一行「单元名称」留空，只填创意列。",
                "配置助手")

    # 固定选项的列做成 Excel 下拉
    for i, c in enumerate(cols, 1):
        f = c["field"]
        if not f:
            continue
        opts = f.get("options")
        if not opts or f.get("match") == "contains":
            continue
        if f.get("type") in ("checkbox_sync_formily", "multiselect_vue", "multiselect_antd"):
            continue        # 多选列不能做单选校验，会把合法值判成非法
        formula = '"' + ",".join(opts) + '"'
        if len(formula) > 255:
            continue
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showErrorMessage=True)
        dv.error = f"只能填：{'/'.join(opts)}"
        ws.add_data_validation(dv)
        col = get_column_letter(i)
        dv.add(f"{col}2:{col}500")

    ws.freeze_panes = "B2"


def build(cfg: dict, positions: list[str], existing_activity: bool = False,
          strategy: dict | None = None, out_name: str | None = None) -> str:
    """按选好的资源位生成模板，返回文件路径。

    existing_activity=True —— 单元挂到一个已经建好的活动上，不生成「活动」sheet
                              （活动ID在界面上填，不进 Excel）
    strategy               —— 策略中心的当前值；不传就读存盘的那套。
                              只用来往「填写说明」页里写一份当前策略的快照 ——
                              人群/内容限制这两组恒定不出列（2026-08-21 起）。
    """
    if not positions:
        raise ValueError("至少要选一个资源位")
    unknown = [p for p in positions if p not in (cfg.get("positions") or {})]
    if unknown:
        raise ValueError(f"配置里没有这些资源位：{unknown}")
    if strategy is None:
        strategy = S.active_payload(cfg)

    wb = Workbook()
    first = wb.active

    if existing_activity:
        wb.remove(first)
    else:
        first.title = "活动"
        _header(first, [{"title": f["name"], "field": f, "role": "unit"}
                        for f in W.columns_for(cfg, positions[0], W.STEP_ACTIVITY)])
        first.freeze_panes = "A2"

    for p in positions:
        ws = wb.create_sheet(sheet_name(p))
        _header(ws, columns_for_sheet(cfg, p))

    _doc_sheet(wb, cfg, positions, existing_activity, strategy)

    name = out_name or f"{cfg['name']}_模板"
    return X.save(wb, f"{name}.xlsx")


def _doc_sheet(wb, cfg: dict, positions: list[str], existing_activity: bool,
               strategy: dict):
    doc = wb.create_sheet("填写说明")
    for col, w in (("A", 26), ("B", 22), ("C", 82)):
        doc.column_dimensions[col].width = w

    rows = [
        ("配置类型", cfg["name"], cfg.get("description", "")),
        ("本次资源位", "、".join(positions), f"共 {len(positions)} 个"),
        ("活动", "挂到已有活动" if existing_activity else "本次新建活动",
         "活动ID在界面上填，模板里没有「活动」sheet" if existing_activity
         else "先填「活动」sheet 那一行，本次所有单元都挂在这个活动下"),
        ("人群 / 内容限制", "跟随策略中心",
         "这两组不在表里 —— 按策略中心配的方案走（同一套 或 按单元名称匹配）"),
        ("", "", ""),
        ("怎么填", "", "① 每个资源位一张「资源位_xxx」表，单元和创意在同一行。"),
        ("", "", "② 一行 = 一个单元 + 它的一条创意；「创意·」开头的列属于创意层。"),
        ("", "", "③ 同一个单元要挂第 2 条创意：重复填写相同的「单元名称」即可；也可紧接着另起一行留空该列，只填创意列。"),
        ("", "", "④ 同一个资源位要投多个单元：填多行，每行都写上不同的「单元名称」。"),
        ("", "", ""),
        ("颜色", "", "黄=必填　灰=选填　橙=条件列（某个字段选了特定值才需要填）　蓝=关联键"),
        ("图片列", "", "填本地图片路径（如 D:\\素材\\banner.png），或直接把图片贴进单元格，两种都行。"),
        ("时间格式", "2026-09-24 10:00:00",
         "年-月-日 时:分:秒。年份不能省，别写「9.24 10:00:00」——"
         "日期控件不认，而且填不进去也不报错，会连累后面几个字段一起填不上。"),
        ("多选列", "", "用英文逗号分隔，例：Android,iPhone"),
        ("", "", ""),
        ("■ 策略中心", "", "下面这些字段不在模板里，跑的时候统一按策略中心配的值填。"),
        ("", "", "要改：回到界面「准备」页 → 打开策略中心。"),
    ]

    for name, val, note in S.summary(cfg, strategy):
        rows.append((f"　{name}", val, note))

    rows.append(("", "", ""))
    for p in positions:
        meta = W.position_meta(cfg, p)
        real = meta.get("real_name")
        rows.append(("", "", ""))
        rows.append((f"■ {p}", f"pos {meta['position_id']}",
                     f"场景 {meta.get('scene','')}　创意系统 {meta.get('system','')}"
                     + (f"　后台实际名称：{real}" if real else "")))
        for c in columns_for_sheet(cfg, p):
            f = c["field"]
            if f is None:
                rows.append((f"　　{c['title']}", "必填", "关联键：留空 = 这一行是上一个单元的又一条创意"))
                continue
            if str(f.get("default", "")).strip():
                tag = "不用填"
            else:
                tag = "必填" if f.get("required") else "选填"
            rows.append((f"　　{c['title']}", tag, W.describe(f)))

    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if str(v).startswith("■") or str(v) in ("怎么填", "颜色"):
                cell.font = Font(bold=True)
