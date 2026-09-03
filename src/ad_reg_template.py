"""常规商广的 Excel 模板：3 列（素材标题 / 素材描述 / 落地页）+ 一页填写说明。

⚠ 只服务 mode: ad_regular。一行 = 一个视频 = 一条创意，行序对应「我的视频」里
  跳过前 K 个之后的第 1、2、3… 个视频。视频数量 N / 跳过前几个 K / 目的·内容·转化
  这些在界面「准备」页填，不进这张表。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import ad_prep as P
from . import ad_reg_data as D
from . import xlsx_kit as X

REQ_FILL = X.FILLS["req"]
OPT_FILL = X.FILLS["opt"]


def build(cfg: dict, prep: dict | None = None, out_name: str | None = None) -> str:
    if prep is None:
        prep = P.load(cfg)

    wb = Workbook()
    ws = wb.active
    ws.title = "素材"

    for i, c in enumerate(D.columns(cfg), 1):
        cell = ws.cell(row=1, column=i, value=c["name"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = REQ_FILL if c.get("required") else OPT_FILL
        ws.column_dimensions[get_column_letter(i)].width = 40 if c["name"] == "素材标题" else 26
        if c.get("_note"):
            cell.comment = Comment(c["_note"], "配置助手")
    ws.freeze_panes = "A2"

    _doc_sheet(wb, cfg, prep)
    name = out_name or f"{cfg['name']}_模板"
    return X.save(wb, f"{name}.xlsx")


def _doc_sheet(wb, cfg: dict, prep: dict):
    doc = wb.create_sheet("填写说明")
    for col, w in (("A", 22), ("B", 20), ("C", 82)):
        doc.column_dimensions[col].width = w

    rows = [
        ("配置类型", cfg["name"], cfg.get("description", "")),
        ("", "", ""),
        ("怎么填", "", "① 一行 = 一个视频 = 一条创意。"),
        ("", "", "② 行序 = 视频顺序：第 1 行配「我的视频」里跳过前 K 个之后的第 1 个视频，"
                 "第 2 行配第 2 个……K 在界面「准备」页填。"),
        ("", "", "③ 每 10 行归一个单元（页面上限 10 条创意/单元）。"),
        ("", "", "④「素材标题」一个单元格里换行写多条，最多 6 条，每条 2~40 字。"),
        ("", "", "⑤ 视频数量、跳过前几个、推广目的/内容、转化目标、出价、投放时间、人群 "
                 "在界面「准备」页填，不进这张表。"),
        ("", "", ""),
        ("颜色", "", "黄=必填"),
        ("", "", ""),
        ("■ 列说明", "", ""),
    ]
    for c in D.columns(cfg):
        rows.append((f"　{c['name']}", "必填" if c.get("required") else "选填", c.get("_note", "")))

    rows += [("", "", ""), ("■ 准备阶段当前的值", "", "改这些请回界面「准备」页")]
    for name, val in P.summary(cfg, prep):
        rows.append((f"　{name}", val, ""))

    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if str(v).startswith("■") or str(v) in ("怎么填", "颜色"):
                cell.font = Font(bold=True)
