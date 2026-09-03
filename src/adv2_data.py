"""原生商广新（三连竞价推广 auto-v2）的数据读取：一张平表 → 三个池子。

⚠ 只服务 mode: ad_v2。老的「原生商广」走 ad_data（计划→单元→创意三层），互不影响。

模板是一张平表，一行 = 一条素材：

    内容            avid              素材标题                    封面
    赘婿 第二季      112868888414555   终于亲上了！！              （贴图或路径）
    凡人修仙传       115390755637733   那就让韩某送道友一程        ...

新页面素材层是「聚合配置」：整个项目就三个池子（稿件≤200 / 标题≤50 / 封面≤100）+
一条描述。所以这里把所有行拍平去重：

  avid    列  → 稿件池（按首次出现去重）
  素材标题 列 → 标题池（按首次出现去重）
  封面    列  → 封面池（留空的行不贡献；同一张图去重）
  描述       → 不在表里，走准备阶段的「素材描述」

⚠ 「内容」列只当人看的备注，不参与任何逻辑。
⚠ 隐藏的行和列直接忽略。封面两种填法都支持：单元格写本地路径 / 图片直接贴在单元格上。
"""
from __future__ import annotations

import logging
import re
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from .wizard_data import DataError, _apply_images, _extract_images

log = logging.getLogger(__name__)

AVID = "avid"
TITLE = "素材标题"
COVER = "封面"

# 页面硬上限（见 docs/三连竞价推广auto-v2-配置项抓取.md §4）
MAX_ARCHIVES = 200
MAX_TITLES = 50
MAX_COVERS = 100
TITLE_MIN, TITLE_MAX = 2, 40


def columns(cfg: dict) -> list[dict]:
    return [dict(c) for c in (cfg.get("columns") or [])]


# ---------------------------------------------------------------- 读表
def _hidden_cols(ws) -> set[int]:
    out: set[int] = set()
    for key, dim in ws.column_dimensions.items():
        if not dim.hidden:
            continue
        try:
            lo = dim.min or column_index_from_string(key)
            hi = dim.max or lo
        except ValueError:
            continue
        out.update(range(int(lo), int(hi) + 1))
    return out


def _hidden_rows(ws) -> set[int]:
    return {int(r) for r, dim in ws.row_dimensions.items() if dim.hidden}


def _rows(path: str) -> tuple[list[dict], dict]:
    """读第一个 sheet 成 [{列名: 值}]，全空行跳过。返回 (行, {列名: 列号})。隐藏行列当不存在。"""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        raise DataError(f"打不开数据文件：{e}") from e

    ws = wb[wb.sheetnames[0]]
    skip_cols = _hidden_cols(ws)
    skip_rows = _hidden_rows(ws)
    if skip_cols or skip_rows:
        log.info("忽略隐藏内容：%d 列、%d 行", len(skip_cols), len(skip_rows))

    headers, idx = [], {}
    for i in range(1, ws.max_column + 1):
        if i in skip_cols:
            headers.append("")
            continue
        v = ws.cell(1, i).value
        h = str(v).strip() if v is not None else ""
        headers.append(h)
        if h:
            idx[h] = i

    out = []
    for r in range(2, ws.max_row + 1):
        if r in skip_rows:
            continue
        rec, empty = {}, True
        for i, h in enumerate(headers, 1):
            if not h:
                continue
            v = ws.cell(r, i).value
            if isinstance(v, float) and v.is_integer():
                v = int(v)
            s = "" if v is None else str(v).strip()
            rec[h] = s
            if s:
                empty = False
        if not empty:
            rec["_row"] = r
            out.append(rec)
    return out, idx


def _dedup(pairs: list[tuple[str, int]]) -> list[dict]:
    """[(值, 行号)] → 按值首次出现去重，返回 [{"value":.., "row":..}]。"""
    seen: set[str] = set()
    out = []
    for val, row in pairs:
        if not val or val in seen:
            continue
        seen.add(val)
        out.append({"value": val, "row": row})
    return out


def load(data_file: str, cfg: dict, settings: dict | None = None) -> dict:
    """读成三个池子：

        {
          "archives": [{"value": "<avid>", "row": n}, ...],   # 去重、保序
          "titles":   [{"value": "<标题>", "row": n}, ...],
          "covers":   [{"value": "<本地路径>", "row": n}, ...],
          "rows": <原始行数>,
        }
    """
    if not data_file:
        raise DataError("还没选数据文件")

    log.info("读数据文件：%s", data_file)
    rows, idx = _rows(data_file)
    if not rows:
        name = Path(data_file).name
        hint = ("　这份是生成出来的空模板（只有表头），请选你自己填过数据的那份表。"
                if "模板" in name else
                "　第 1 行是表头，数据从第 2 行开始；整行被隐藏的也会被跳过。")
        raise DataError(f"「{name}」里一行数据都没有。\n{hint}\n完整路径：{data_file}")

    missing = [c["name"] for c in columns(cfg)
               if c.get("required") and c["name"] not in idx]
    if missing:
        raise DataError(f"「{Path(data_file).name}」表头缺少必填列：{'、'.join(missing)}。\n"
                        f"　现有列：{'、'.join(k for k in idx) or '（空）'}"
                        f"　—— 整列被隐藏的列也算不存在。")

    imgs = _extract_images(data_file)
    sheet = next(iter(imgs), "")
    if imgs:
        _apply_images(rows, idx, imgs.get(sheet, {}), sheet)

    archives = _dedup([(str(r.get(AVID, "")).strip(), r["_row"]) for r in rows])
    titles = _dedup([(str(r.get(TITLE, "")).strip(), r["_row"]) for r in rows])
    covers = _dedup([(str(r.get(COVER, "")).strip(), r["_row"]) for r in rows])

    return {"archives": archives, "titles": titles, "covers": covers, "rows": len(rows)}


def validate(cfg: dict, data: dict) -> list[str]:
    """跑之前的体检。返回人话的问题清单。"""
    issues = []
    arch, titles, covers = data["archives"], data["titles"], data["covers"]

    if not arch:
        issues.append("一个 avid 都没有 —— 稿件池不能是空的")
    for a in arch:
        if not re.fullmatch(r"\d+", a["value"]):
            issues.append(f"第{a['row']}行的 avid「{a['value']}」不是纯数字")
    if len(arch) > MAX_ARCHIVES:
        issues.append(f"去重后有 {len(arch)} 个 avid，超过页面上限 {MAX_ARCHIVES}")

    if not titles:
        issues.append("一个素材标题都没有 —— 标题池不能是空的")
    for t in titles:
        n = len(t["value"])
        if n < TITLE_MIN or n > TITLE_MAX:
            issues.append(f"第{t['row']}行的素材标题 {n} 字，需 {TITLE_MIN}~{TITLE_MAX} 字：{t['value']}")
    if len(titles) > MAX_TITLES:
        issues.append(f"去重后有 {len(titles)} 个素材标题，超过页面上限 {MAX_TITLES}。"
                      f"先合并重复 / 相近的标题")

    if len(covers) > MAX_COVERS:
        issues.append(f"去重后有 {len(covers)} 张封面，超过页面上限 {MAX_COVERS}")
    for c in covers:
        p = c["value"]
        if ("/" in p or "\\" in p) and not Path(p).exists():
            issues.append(f"第{c['row']}行的封面文件不存在：{p}")

    return issues
