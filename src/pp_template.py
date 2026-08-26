"""价格面板配置的 Excel 模板。

⚠ 只服务 mode: price_panel。

产出的表：
  「活动」    一行，活动层字段；**挂到已有活动时不生成这张表**（和资源位投放一样）
  「单元」    一行一个单元，列 = unit_fields + excel_from_unit + 按 SKU 展开的列
  「填写说明」每列什么意思 + 策略中心现在配的是什么

单元表的列分四段：
  单元层        unit_fields（名字/时间/三段面板）+ excel_from_unit 借来的页面字段
                （生效渠道、人群、收银台类型、选中类型、赠单片）
  单元层·按SKU  「<SKU>·<字段>」，上了面板才出（目前只有异形SKU 的 pid / 搭售商品ID）
  创意层·按SKU  「<SKU>·<字段>」，字段随这个 SKU 的搭售类型变
  面板级        价格面板切换按钮，只有配了面板2 才出

⚠ 后三段**不枚举 26 个 SKU** —— 只出这次真用得到的。哪些 SKU 上面板、哪个 SKU
搭什么，都由策略中心「SKU选择 + 面板套餐 + 套餐排序」那一组推出来
（见 pp_data.sku_columns）。换了方案要重新生成一次模板，列才跟着变。

人群 2026-08-26 起改成 Excel 里逐单元填（一批单元里「投给谁」几乎每行都不一样）。
"""
from __future__ import annotations

import logging
import shutil
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import pp_data as D
from . import wizard_strategy as S
from .paths import resource, user_path

log = logging.getLogger(__name__)

REQ_FILL = PatternFill("solid", fgColor="FFF2CC")     # 必填 浅黄
OPT_FILL = PatternFill("solid", fgColor="F2F2F2")     # 选填 浅灰
KEY_FILL = PatternFill("solid", fgColor="DDEBF7")     # 关联键 浅蓝
CRE_FILL = PatternFill("solid", fgColor="E2EFDA")     # 创意层（按 SKU）浅绿
SKU_FILL = PatternFill("solid", fgColor="FFF2E8")     # 单元层（按 SKU）浅橙


def _describe(f: dict) -> str:
    bits = []
    sk = f.get("_sku")
    if sk and sk.get("sku"):
        bits.append(f"{sk['layer']} · {sk['sku']}（搭售：{sk['tie']}）")
    elif sk:
        bits.append("创意层 · 面板级")
    if f.get("_note"):
        bits.append(f["_note"])
    if f.get("options"):
        bits.append("可选：" + "、".join(str(o) for o in f["options"]))
    if f.get("max"):
        bits.append(f"最多 {f['max']} 字")
    if f.get("size"):
        bits.append(f"尺寸 {f['size']}")
    if str(f.get("type", "")).startswith("pp_upload"):
        bits.append("填本地路径 / 图片网址，或者直接把图贴进这一格")
    if f.get("type") == "pp_fill_or_upload":
        bits.append("填链接，或者给本地文件路径走上传")
    when = f.get("_when")
    if when:
        bits.append(f"只在「{when[0]}」= {'/'.join(str(v) for v in when[1])} 时才填")
    if str(f.get("default", "")).strip():
        bits.append(f"留空就用：{f['default']}")
    if f.get("required"):
        bits.append("必填")
    return "；".join(bits)


def build(cfg: dict, existing_activity: bool = False) -> str:
    """产出两个文件，返回值是主模板的路径。

    existing_activity=True（界面上选了「挂到已有活动」）就不生成「活动」sheet ——
    活动ID 在界面上填，Excel 里不该再有第二个地方能填它。

    ⚠ PID 映射表是**单独一个文件**：它跟着「这一批 pid 是什么」走，
      换一次投放通常不用重做，和逐次都要重填的单元表不是一个生命周期。
      路径在策略中心「PID映射」那一组里指定。
    """
    payload = S.active_payload(cfg)
    per_sku = D.sku_columns(cfg, payload)
    channel = D.channel_of(cfg)

    wb = Workbook()
    if existing_activity:
        _units_sheet(wb.active, cfg, per_sku, channel)
    else:
        _activity_sheet(wb.active, cfg)
        _units_sheet(wb.create_sheet("单元"), cfg, per_sku, channel)
    _help_sheet(wb.create_sheet("填写说明"), cfg, per_sku, existing_activity, channel)
    path = user_path("data", f"{cfg['name']}_模板.xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    _build_pid(cfg)
    return str(path)


def pid_path(cfg: dict):
    return user_path("data", f"{cfg['name']}_PID映射表.xlsx")


def _build_pid(cfg: dict) -> str:
    """PID 映射表。

    优先**把 assets 里那份默认表拷过来** —— 那是运营一格一格抄出来的真 pid
    （5 套方案、190 多行），新装的人开箱就能用，不用再抄一遍。
    assets 里没有才退回去生成一张空表（表头 + 每个 SKU 一行）。

    ⚠ 已经有这个文件就不覆盖 —— 它是人自己维护的，重生成一次模板就把人家的 pid
      冲掉，这种事只要发生一次就再没人敢点「生成模板」。
    """
    p = pid_path(cfg)
    if p.exists():
        return str(p)

    default = resource("assets", f"{cfg['name']}_PID映射表.xlsx")
    if default and Path(default).exists():
        p.parent.mkdir(parents=True, exist_ok=True)
        shutil.copyfile(default, p)
        log.info("PID 映射表用了自带的默认表：%s", default)
        return str(p)
    cols = D.pid_columns(cfg)
    wb = Workbook()
    ws = wb.active
    ws.title = D.pid_sheet(cfg)
    for i, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        c.fill = KEY_FILL if name in ("方案名", "SKU") else OPT_FILL
        c.comment = Comment(_PID_NOTES.get(name, ""), "配置助手") if _PID_NOTES.get(name) else None
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(30, len(name) * 2.6))
    skus = [s for s in D.sku_list(cfg) if s not in (cfg.get("sku_map_skip") or [])]
    for r, sku in enumerate(skus, start=2):
        ws.cell(row=r, column=cols.index("方案名") + 1, value="默认")
        ws.cell(row=r, column=cols.index("SKU") + 1, value=sku)
    ws.freeze_panes = "C2"
    _pid_help(wb.create_sheet("填写说明"), cfg)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    return str(p)


_PID_NOTES = {
    "方案名": "同一份表里可以放好几套，用这一列区分；策略中心的「PID映射方案」选用哪一套。"
              "只有一套的话，这一列可以整列留空。",
    "SKU": "页面上的卡种名，必须一字不差。异形SKU 不在这里配（它在 Excel 里逐单元填）。",
    "价格面板pid": "只填 pid 数字，例 11439。这个 SKU 被标了「买赠」时必填。",
    "买赠商品类型": "观影券 / 会员天数 / 永久装扮 / 普通装扮 / 会员购抽奖机会 / "
                    "会员购运费券 / 会员购优惠券 / 会员购魔晶 / 道具",
    "买赠商品ID": "选完商品类型才有这一栏，只填 id 数字。",
    "组合价格": "这个 SKU 被标了「0元购」时必填。每组写成「单会员价格pid:加购商品id」，"
                "多组用英文逗号分隔。",
}


def _pid_help(ws, cfg: dict):
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 86
    rows = [
        ("这张表是干嘛的", "只回答「搭什么」：每个 SKU 配哪个 pid、买赠什么商品、0元购加购什么。"
                          "页面上 pid 的下拉一个 SKU 就有一两百条，选哪个只能人定，"
                          "所以在这里一次抄好、整批套用。"),
        ("它不管什么", "⚠ 不管「哪些 SKU 搭售」。那是在策略中心的套餐卡片上点角标定的"
                      "（无 / 买赠 / 0元购 / 买赠+0元购）。2026-08-26 起这张表没有「搭售类型」那一列了 —— "
                      "两处都能说同一件事的话，不一致时谁赢没人讲得清。"),
        ("怎么用", "在程序的「策略中心 → PID映射」里把这个文件的路径指过去（那儿有个「浏览…」），"
                  "再在套餐那一组里选「PID映射方案」——可以全部单元用同一套，也可以按单元名称关键词切。"),
        ("一套 = 一批行", "「方案名」相同的行算一套。新客面板和老客面板用不同 pid，就写两套。"),
        ("不用的 SKU", "整行删掉，或者留着不填都行 —— 只有这次真被标了搭售的 SKU 才会被检查。"),
        ("异形SKU", "不在这张表里。它页面上没有「搭售类型」这一档，pid 和搭售商品ID 在单元 Excel 里逐单元填。"),
        ("重新生成模板", "不会覆盖这个文件。要换一份就先改名或删掉。"),
    ]
    for r, (a, b) in enumerate(rows, 1):
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")


def _activity_sheet(ws, cfg: dict):
    """活动层：一行。本批所有单元都挂在它下面。"""
    ws.title = D.activity_sheet(cfg)
    for i, f in enumerate(D.activity_fields(cfg), 1):
        cell = ws.cell(row=1, column=i, value=f["name"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = REQ_FILL if f.get("required") else OPT_FILL
        note = _describe(f)
        if note:
            cell.comment = Comment(note, "配置助手")
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(34, len(f["name"]) * 2.6))
    ws.freeze_panes = "A2"


def _units_sheet(ws, cfg: dict, per_sku: list[dict] | None = None, channel: str = "全局"):
    ws.title = "单元"
    fields = D.unit_fields(cfg, channel) + list(per_sku or [])
    for i, f in enumerate(fields, 1):
        cell = ws.cell(row=1, column=i, value=f["name"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        if f["name"] == D.UNIT_NAME:
            cell.fill = KEY_FILL
        elif (f.get("_sku") or {}).get("layer") == "创意层":
            cell.fill = CRE_FILL
        elif f.get("_sku"):
            cell.fill = SKU_FILL
        elif f.get("required"):
            cell.fill = REQ_FILL
        else:
            cell.fill = OPT_FILL
        note = _describe(f)
        if note:
            cell.comment = Comment(note, "配置助手")
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(34, len(f["name"]) * 2.6))
        # ⚠ 时间列先把格式设成时间：用户敲「2026-09-24 10:00」时 Excel 存成真正的
        #   时间值，读出来就是标准写法。存成文本的话「9.24 10:00」也能填进去，
        #   跑到页面上才发现日期控件不认。
        if f.get("type") == "pp_date":
            for r in range(2, 301):
                ws.cell(row=r, column=i).number_format = "yyyy-mm-dd hh:mm:ss"
    ws.freeze_panes = "A2"


def _help_sheet(ws, cfg: dict, per_sku: list[dict] | None = None,
                existing_activity: bool = False, channel: str = "全局"):
    per_sku = list(per_sku or [])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 78
    r = 1

    def head(text):
        nonlocal r
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, size=12)
        r += 1

    def line(a, b="", c=""):
        nonlocal r
        ws.cell(row=r, column=1, value=a)
        ws.cell(row=r, column=2, value=b)
        cell = ws.cell(row=r, column=3, value=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1

    head("这张表怎么填")
    line("一行 = 一个单元", "", "「单元」表里一行一个单元；空行和没填单元名称的行会被跳过")
    if existing_activity:
        line("活动", "", "这份模板选的是「挂到已有活动」，所以没有「活动」sheet —— "
                        "活动ID 在程序的「投放配置」页上填，本批单元都挂到那个活动下。")
    else:
        line("活动", "", "「活动」sheet 填一行，本批所有单元都挂在这一个活动下（跑的时候先建活动、再建单元）。"
                        "要挂到已有活动的话，回「投放配置」页把活动切成「挂到已有活动」，重新生成一次模板。")
    line("SKU 放哪几个、怎么排", "",
         "不在这张表里 —— 在程序的「策略中心」里，那一组叫「SKU选择 + 面板套餐 + 套餐排序」，"
         "按顺序点选就行，点的顺序就是页面上从左到右的顺序。"
         "面板个数也不用管：面板2 一个都不选就是 1个，选了就是 2个，"
         "再选上「隐藏sku面板」就是「2个+查看更多」。"
         "搭售也在这一组：点套餐卡片右边那个小角标，四选一（无 / 买赠 / 0元购 / 买赠+0元购）；"
         "搭什么、用哪个 pid 则由「PID映射方案」指到 PID 映射表里的一套。")
    line("人群", "", "在这张表里逐单元填（2026-08-26 从策略中心搬过来的）。"
                    "「人群选组」选「不限」就不用管后面几列；选了别的，才填它下面那一列具体投谁。")
    line("按 SKU 的那些列", "", "表头带「<SKU>·」的：橙色是单元层（异形SKU 的 pid / 搭售商品ID），"
                              "绿色是创意层。⚠ 只出这次真用得到的 SKU —— 由策略中心那组方案推出来的"
                              "（上了面板的 SKU 才有列，角标标了买赠的再多出买赠那几列）。"
                              "改了套餐或角标，要重新生成一次模板列才跟着变。")
    r += 1

    if not existing_activity:
        head("「活动」sheet 的列")
        for f in D.activity_fields(cfg):
            line(f["name"], "必填" if f.get("required") else "", _describe(f))
        r += 1

    head("「单元」表 · 单元层的列")
    if channel == D.DIRECT:
        line("生效渠道 = 定向", "", "页面上这一套只有 18 个字段：人群、生效平台、运营商、省份、"
                                  "投放区域、频次、优先级、内容设置、赠单片、创意赛马**都不存在**，"
                                  "所以这张表里也不出这些列。要它们就回「投放配置」页把生效渠道切回「全局」，"
                                  "再重新生成一次模板。"
                                  "另：策略中心的「生效平台」定向下仍然有用 —— 每个卡种填哪几个 pid "
                                  "就是按它去 PID 映射表里挑的，留空才是「这个卡种有几个就填几个」。")
    for f in D.unit_fields(cfg, channel):
        line(f["name"], "必填" if f.get("required") else "", _describe(f))
    r += 1

    unit_sku = [f for f in per_sku if (f.get("_sku") or {}).get("layer") == "单元层"]
    cre_sku = [f for f in per_sku if (f.get("_sku") or {}).get("layer") != "单元层"]

    if unit_sku:
        head("「单元」表 · 单元层里按 SKU 展开的列")
        line("怎么来的", "", "这几列填的是**单元页**上那张 SKU 卡片，不是创意页。"
                            "目前只有异形SKU：它页面上没有「搭售类型」这一档，就固定一对 "
                            "价格面板pid + 搭售商品ID，而且一个单元一个样，所以只能逐单元填。")
        for f in unit_sku:
            line(f["name"], "必填" if f.get("required") else "", _describe(f))
        r += 1

    if cre_sku:
        head("「单元」表 · 创意层的列")
        line("怎么来的", "", "上了面板的每个 SKU 都有「角标文案 / 红包弹窗动画」两列；"
                            "角标标成「买赠」的 SKU 再多出 配置类型 / 小灰条文案 / icon / 商品权益名称 / "
                            "支付成功商品图 / 权益跳转链接。"
                            "「0元购」在创意页上只多一个只读回显的「加购商品id」，没有要填的，所以不出列。")
        for f in cre_sku:
            line(f["name"], "必填" if f.get("required") else "", _describe(f))
        r += 1
    else:
        head("「单元」表 · 创意层的列")
        line("一列都没有", "", "策略中心的「SKU选择 + 面板套餐 + 套餐排序」还没配 —— "
                             "面板里放了哪些 SKU 都不知道，就没法知道该出哪些列。"
                             "去配一套方案，再重新生成一次模板。")
        r += 1

    head("策略中心现在配的是什么")
    payload = S.active_payload(cfg)
    rows = S.summary(cfg, payload)
    if not rows:
        line("（还没配）", "", "打开程序的「策略中心」配一次，这里就会列出来")
    for a, b, c in rows:
        line(a, b, c)
