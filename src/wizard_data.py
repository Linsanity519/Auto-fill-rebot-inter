"""wizard 模式的数据读取：把模板读成「活动 → 单元 → 创意」嵌套结构。

⚠ 只服务 mode: wizard。老配置走 src/datasource.py，两边互不影响。

模板结构（见 wizard_template）：单元和创意合在同一张「资源位_xxx」表里，
一行 = 一个单元 + 一条创意；「单元名称」留空的行，是上一个单元的又一条创意。
创意层的列带「创意·」前缀，读的时候在这里剥掉。

三个来源合成一条单元数据：
  · Excel 行          —— 单元名称、投放时间、优先级这些逐单元不同的
  · 策略中心          —— 生效平台/投放流量池/创意赛马…… 见 wizard_strategy
  · 界面上的活动设置  —— 挂到已有活动时，活动ID不进 Excel，从 settings 传进来

图片列有三种填法，都支持：
  · 单元格里写本地路径      → 直接用
  · 图片直接贴在单元格上    → 从 xlsx 里抽出来存到 output/_images/ 再用
  · http(s) 网址            → 执行时下载（见 src/images.py）

⚠ yaml 里带 default 的列（试看付费条的 IP 图）在这里补：留空就用固定值。
  补在读取阶段，是为了让后面的校验、核对页、filler 看到的都是同一份数据。
"""
from __future__ import annotations

import logging
import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from . import wizard_schema as W
from .images import is_url
from . import wizard_strategy as S
from . import wizard_template as WT
from .paths import user_path

log = logging.getLogger(__name__)

UNIT_NAME = "单元名称"
CREATIVE_PREFIX = WT.CREATIVE_PREFIX
IMG_DIR = "_images"

# 旧模板（单元_xxx / 创意_xxx 两张表）已经不认了，碰到就提示重新生成
OLD_SHEET_PREFIXES = ("单元_", "创意_")


class DataError(Exception):
    pass


def _sheet_rows(wb, title: str) -> list[dict]:
    """读一个 sheet 成 [{列名: 值}]，全空行跳过。"""
    if title not in wb.sheetnames:
        raise DataError(f"模板里没有 sheet「{title}」")
    ws = wb[title]
    headers = []
    for i in range(1, ws.max_column + 1):
        v = ws.cell(1, i).value
        headers.append(str(v).strip() if v is not None else "")

    rows = []
    for r in range(2, ws.max_row + 1):
        rec, empty = {}, True
        for i, h in enumerate(headers, 1):
            if not h:
                continue
            v = ws.cell(r, i).value
            s = "" if v is None else str(v).strip()
            rec[h] = s
            if s:
                empty = False
        if not empty:
            rec["_row"] = r
            rows.append(rec)
    return rows


# WPS「嵌入单元格」的图片：单元格里是公式 =DISPIMG("ID_xxx",1)，
# 图片本体在 xl/cellimages.xml，openpyxl 的 ws._images 读不到。
_DISPIMG_RE = re.compile(r'DISPIMG\(\s*"([^"]+)"', re.IGNORECASE)


def _local(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _sniff_ext(data: bytes, fallback: str) -> str:
    """按文件头认扩展名。后台上传框按扩展名判类型，名不副实的会被拒。"""
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return ".png"
    if data[:3] == b"\xff\xd8\xff":
        return ".jpg"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return ".gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return ".webp"
    ext = ("." + fallback.rsplit(".", 1)[-1].lower()) if "." in fallback else ""
    return ".jpg" if ext == ".jpeg" else (ext or ".png")


def _wps_cell_images(path: str) -> dict[str, tuple[bytes, str]]:
    """读 WPS 嵌入单元格图片：返回 {'ID_xxx': (图片字节, '.png')}。

    结构：xl/cellimages.xml 里每个 <xdr:pic> 带 <xdr:cNvPr name="ID_xxx">
    和 <a:blip r:embed="rIdN">；rIdN → 媒体文件在 xl/_rels/cellimages.xml.rels。
    """
    out: dict[str, tuple[bytes, str]] = {}
    try:
        with zipfile.ZipFile(path) as z:
            names = set(z.namelist())
            if "xl/cellimages.xml" not in names:
                return out
            rels: dict[str, str] = {}
            if "xl/_rels/cellimages.xml.rels" in names:
                rroot = ET.fromstring(z.read("xl/_rels/cellimages.xml.rels"))
                for rel in rroot:
                    rid = rel.get("Id")
                    tgt = rel.get("Target") or ""
                    if rid and tgt:
                        rels[rid] = tgt.lstrip("/").replace("xl/", "", 1) if tgt.startswith("/xl/") else tgt

            root = ET.fromstring(z.read("xl/cellimages.xml"))
            for pic in root.iter():
                if _local(pic.tag) != "pic":
                    continue
                name = embed = None
                for node in pic.iter():
                    ln = _local(node.tag)
                    if ln == "cNvPr" and node.get("name"):
                        name = node.get("name")
                    elif ln == "blip":
                        for k, v in node.attrib.items():
                            if _local(k) == "embed":
                                embed = v
                if not (name and embed and embed in rels):
                    continue
                target = rels[embed]
                for cand in (f"xl/{target}", target, f"xl/media/{target.rsplit('/', 1)[-1]}"):
                    if cand in names:
                        data = z.read(cand)
                        out[name] = (data, _sniff_ext(data, cand))
                        break
    except Exception as e:
        log.warning("读 WPS 嵌入单元格图片失败，忽略：%s", e)
    return out


def _extract_images(path: str) -> dict[str, dict[str, str]]:
    """抽出贴在单元格里的图片。

    返回 {sheet名: {'B3': 存好的文件路径}}。
    openpyxl 读图需要 rich text 关闭，这里单独开一次 workbook 取 _images。
    同时处理 WPS 的「嵌入单元格」图片（单元格里是 =DISPIMG("ID_xxx",1) 公式）。
    """
    out: dict[str, dict[str, str]] = {}
    try:
        wb = load_workbook(path)
    except Exception as e:
        log.warning("读图片失败，忽略：%s", e)
        return out

    img_dir = user_path("output", IMG_DIR)

    # ---- WPS 嵌入单元格图片：按 sheet 扫 =DISPIMG("ID_xxx") 公式，落到对应单元格 ----
    wps = _wps_cell_images(path)
    if wps:
        for ws in wb.worksheets:
            cells: dict[str, str] = {}
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or "DISPIMG" not in v.upper():
                        continue
                    m = _DISPIMG_RE.search(v)
                    if not m or m.group(1) not in wps:
                        continue
                    try:
                        img_dir.mkdir(parents=True, exist_ok=True)
                        data, ext = wps[m.group(1)]
                        dst = img_dir / f"{ws.title}_{cell.coordinate}_dispimg{ext}"
                        dst.write_bytes(data)
                        cells[cell.coordinate] = str(dst)
                    except Exception as e:
                        log.warning("落 WPS 图片 %s!%s 失败：%s", ws.title, cell.coordinate, e)
            if cells:
                out[ws.title] = cells

    for ws in wb.worksheets:
        imgs = getattr(ws, "_images", None) or []
        if not imgs:
            continue
        cells: dict[str, str] = {}
        for n, im in enumerate(imgs):
            anchor = getattr(im, "anchor", None)
            frm = getattr(anchor, "_from", None)
            if frm is None:
                continue
            ref = f"{get_column_letter(frm.col + 1)}{frm.row + 1}"
            try:
                data = im._data() if callable(getattr(im, "_data", None)) else None
                if not data:
                    continue
                img_dir.mkdir(parents=True, exist_ok=True)
                ext = (getattr(im, "format", None) or "png").lower()
                dst = img_dir / f"{ws.title}_{ref}_{n}.{ext}"
                dst.write_bytes(data)
                cells[ref] = str(dst)
            except Exception as e:
                log.warning("抽图 %s!%s 失败：%s", ws.title, ref, e)
        if cells:
            out.setdefault(ws.title, {}).update(cells)
    return out


def _is_dispimg(val: str) -> bool:
    return "DISPIMG(" in str(val).upper().replace(" ", "")


def _apply_images(rows: list[dict], headers_idx: dict[str, int],
                  imgs: dict[str, str], sheet: str):
    """把抽出来的图片路径填回对应单元格的值。

    单元格原本为空、或里面是 WPS 的 =DISPIMG(...) 公式（读出来就是那串文本，
    不是图）时才覆盖 —— 后者必须换掉，不然当成本地路径去找一定报「图片找不到」。
    """
    if not imgs:
        return
    for rec in rows:
        r = rec.get("_row")
        for col_name, col_i in headers_idx.items():
            ref = f"{get_column_letter(col_i)}{r}"
            cur = rec.get(col_name)
            if ref in imgs and (not cur or _is_dispimg(cur)):
                rec[col_name] = imgs[ref]
                log.info("%s!%s 用了贴在单元格里的图片", sheet, ref)


def load(path: str, cfg: dict, settings: dict | None = None) -> dict:
    """读模板，返回 {activity: {...}, units: [{position, header, creatives:[...]}]}

    settings 里认这两个键（webapp 的「准备」页填的，命令行/tk 版没有就走默认）：
      wizard_activity  {"existing": bool, "activity_id": str, "activity_type": str}
      wizard_strategy  策略中心当前那套；不给就读存盘的
    """
    settings = settings or {}
    p = Path(path or "")
    if not p.exists():
        raise DataError(f"数据文件不存在：{p.resolve()}")

    wb = load_workbook(p, data_only=True)
    imgs = _extract_images(str(p))
    strategy = settings.get("wizard_strategy") or S.active_payload(cfg)

    activity = _activity(wb, settings)
    # 活动层也认 default（现在一个都没有，但别留这个坑：加了 default 却不生效最难查）
    W.apply_defaults(activity, W.defaults_for(cfg, W.position_names(cfg)[0], W.STEP_ACTIVITY))

    # ---- 各资源位 ----
    units = []
    used: list[str] = []
    supplied: dict[str, list[str]] = {}   # 这个资源位有哪些字段是策略中心供的
    extra: list[str] = []                 # 读的时候就发现的问题（人群预设撞车之类）
    for pos in W.position_names(cfg):
        title = WT.sheet_name(pos)
        if title not in wb.sheetnames:
            continue                     # 这次没选这个资源位
        got, names, notes = _units_of(wb, title, pos, cfg, strategy, imgs.get(title, {}))
        if not got:
            continue        # 表在但一行没填：当没选这个资源位，别拿它的策略去卡人
        used.append(pos)
        units += got
        supplied[pos] = names
        extra += notes

    if not units:
        if any(s.startswith(OLD_SHEET_PREFIXES) for s in wb.sheetnames):
            raise DataError(
                "这份模板是旧版格式（单元/创意分成两张表）。"
                "现在单元和创意合并成一张「资源位_xxx」表了，请重新生成模板再填。")
        raise DataError("没读到任何单元数据。确认模板里「资源位_xxx」表填了内容")

    return {"activity": activity, "units": units, "positions": used,
            "strategy": strategy, "supplied": supplied, "extra_issues": extra}


def _activity(wb, settings: dict) -> dict:
    """活动信息：界面上勾了「挂到已有活动」就用界面填的，否则读「活动」sheet。

    ⚠ 返回值里的「已有活动ID」是 runner 的分叉点：有值 = 不新建活动，
      直接往那个活动下加单元。这里统一成这一个键，runner 不用关心值从哪来。
    """
    act_cfg = settings.get("wizard_activity") or {}
    if act_cfg.get("existing"):
        aid = str(act_cfg.get("activity_id", "")).strip()
        if not aid:
            raise DataError("选了「挂到已有活动」，但没填活动ID")
        # 活动类型ID 界面上不问了：直连单元页的 URL 里要带一个，
        # 填 5（测试验收）对新建单元没有影响，需要指定时用命令行 --activity-type
        return {"已有活动ID": aid,
                "活动类型ID": str(act_cfg.get("activity_type", "") or "5").strip(),
                "活动名称": str(act_cfg.get("activity_name", "")).strip()}

    if "活动" not in wb.sheetnames:
        raise DataError(
            "模板里没有「活动」sheet。要挂到已有活动的话，在「准备」页把活动切到"
            "「挂到已有活动」并填活动ID；要新建活动就重新生成一份带活动 sheet 的模板。")
    act_rows = _sheet_rows(wb, "活动")
    if not act_rows:
        raise DataError("「活动」sheet 里没有数据")
    if len(act_rows) > 1:
        raise DataError(f"「活动」sheet 有 {len(act_rows)} 行，只能填一行")
    return act_rows[0]


def _units_of(wb, title: str, pos: str, cfg: dict, strategy: dict,
              imgs: dict[str, str]) -> tuple[list[dict], list[str], list[str]]:
    """一张「资源位_xxx」表 → (若干单元, 这张表里由策略中心补的字段名, 读到的问题)。

    ⚠ 表里有的列以表为准，表里没有的才拿策略中心的。人群/内容限制现在恒定不出列，
      但老模板里可能还留着，靠这条规则自动适配，不用再传一个开关进来。
    """
    ws = wb[title]
    headers = {h: i for i, h in enumerate(_headers_of(ws), 1) if h}
    rows = _sheet_rows(wb, title)
    _apply_images(rows, headers, imgs, title)
    unit_defaults = W.defaults_for(cfg, pos, W.STEP_UNIT)
    creative_defaults = W.defaults_for(cfg, pos, W.STEP_CREATIVE)

    # 这张表里「该由策略中心供」的字段：这个资源位有、模板里又没这一列的。
    # ⚠ 按「该不该供」算，不是按「供上了没有」算 —— 策略里漏配的必填项要报成
    #   「策略中心没配」，报成「第N行没填」会让人去 Excel 里找一个根本不存在的列。
    have = W.unit_field_names(cfg, pos)
    responsible = [n for n in W.strategy_names(cfg) if n in have and n not in headers]
    base = {k: v for k, v in S.resolve(cfg, strategy, pos).items() if k in responsible}
    out: list[dict] = []
    notes: list[str] = []
    by_name: dict[str, dict] = {}        # 单元名 → 那个单元，同名的创意往里追加
    current_unit: dict | None = None      # 当前行所属单元，供下一行留空名称时续写
    for rec in rows:
        name = str(rec.get(UNIT_NAME, "")).strip()
        creative = {k[len(CREATIVE_PREFIX):]: v
                    for k, v in rec.items() if k.startswith(CREATIVE_PREFIX)}
        for f in W.apply_defaults(creative, creative_defaults):
            log.info("「%s」第%s行 创意·%s 没填，用固定值", title, rec["_row"], f)
        if not name:
            # 「单元名称」留空 = 紧上一行所属单元的又一条创意。
            # 不能用 out[-1]：同名单元可能在中间又出现一次，此时它并不一定
            # 是首次出现顺序里的最后一个单元。
            if current_unit is None:
                raise DataError(
                    f"「{title}」第{rec['_row']}行没填单元名称，"
                    f"上面也没有可以挂的单元（留空只在续写上一个单元的创意时才成立）")
            current_unit["creatives"].append(creative)
            continue

        # 人群/内容限制按单元名称匹配时，每一行可能不一样，得按这一行的名字再解一次
        note = S.notes(cfg, strategy, pos, name)
        strat = base
        if note:
            strat = {k: v for k, v in S.resolve(cfg, strategy, pos, name).items()
                     if k in responsible}

        if name in by_name:
            # 同名 = 同一个单元的又一条创意。留空续写是简写，写同一个名字更直观，
            # 两种都认（运营两种都会用）。
            current_unit = by_name[name]
            current_unit["creatives"].append(creative)
            continue

        header = {k: v for k, v in rec.items() if not k.startswith(CREATIVE_PREFIX)}
        header.update(strat)             # 策略中心接管的字段，模板里没有这些列
        W.apply_defaults(header, unit_defaults)
        unit = {"position": pos, "header": header, "creatives": [creative],
                "row": rec["_row"], "strategy_note": note}
        by_name[name] = unit
        out.append(unit)
        current_unit = unit
    return out, responsible, notes


def _headers_of(ws) -> list[str]:
    return [(str(ws.cell(1, i).value).strip() if ws.cell(1, i).value is not None else "")
            for i in range(1, ws.max_column + 1)]


# ---------------------------------------------------------------- 校验
def validate(cfg: dict, data: dict) -> list[str]:
    """离线校验，返回问题清单。空 = 通过。"""
    issues: list[str] = []

    # 填了「已有活动ID」= 挂到现成活动，活动层字段一个都不用填
    if not str(data["activity"].get("已有活动ID", "")).strip():
        act_fields = W.columns_for(cfg, W.position_names(cfg)[0], W.STEP_ACTIVITY)
        issues += _check_row(act_fields, data["activity"], "活动")

    supplied = data.get("supplied") or {}
    issues += strategy_issues(cfg, data.get("strategy"), supplied)
    issues += list(data.get("extra_issues") or [])

    for u in data["units"]:
        pos = u["position"]
        label = f"[{pos}] 第{u['row']}行"
        name = str(u["header"].get(UNIT_NAME, "")).strip()

        # 策略中心供的字段单独在 strategy_issues 里报一次，不跟着每一行重复喊；
        # 但「按单元名称匹配」模式下人群是逐行不同的，只能在这里按行查
        skip = set(supplied.get(pos, [])) - _per_unit_names(cfg, data.get("strategy"))
        unit_fields = [f for f in W.columns_for(cfg, pos, W.STEP_UNIT) if f["name"] not in skip]
        issues += _check_row(unit_fields, u["header"], label)
        issues += _scheme_hints(cfg, data.get("strategy"), label, name)
        for msg in S.merge_conflicts(cfg, data.get("strategy"), pos, name):
            issues.append(f"{label}：{msg}")

        for i, c in enumerate(u["creatives"], 1):
            issues += _check_row(W.columns_for(cfg, pos, W.STEP_CREATIVE), c,
                                 f"{label} 创意{i}")
    return issues


def _per_unit_names(cfg: dict, strategy: dict | None) -> set:
    """「按单元名称匹配」模式下那几组的字段是逐行不同的，不能按资源位一次性查。"""
    return S.per_unit_names(cfg, strategy)


def _scheme_hints(cfg: dict, strategy: dict | None, label: str, name: str) -> list[str]:
    """方案没落上时，把话说到点子上：是名字没命中，还是方案本身没配。

    ⚠ 不这么区分的话，只会看到一句「策略中心没配人群选组」——
      而实际上策略配得好好的，是这个单元的名字一个关键词都不含。
    """
    if not name:
        return []
    return [f"{label}：{msg}" for msg in S.unmatched_hint(cfg, strategy, name)]


def strategy_issues(cfg: dict, strategy: dict | None, supplied: dict) -> list[str]:
    """策略中心的值够不够、合不合法。按资源位报一次，不跟着单元行重复。

    supplied = {资源位: [这张表里由策略中心补的字段名]}，只查真正用到的那些
    ——人群改成逐单元填之后，策略里那几个人群字段没配也不该拦人。
    """
    out: list[str] = []
    per_unit = _per_unit_names(cfg, strategy)
    for pos, names in (supplied or {}).items():
        names = set(names) - per_unit
        fields = [f for f in W.columns_for(cfg, pos, W.STEP_UNIT) if f["name"] in names]
        if not fields:
            continue
        resolved = S.resolve(cfg, strategy, pos)
        for msg in _check_row(fields, resolved, f"[策略中心] {pos}"):
            if "必填但为空" in msg:
                msg += "（在「准备」页打开策略中心配置）"
            out.append(msg)
    return out


TIME_FORMATS = ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
                "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d")


def _looks_like_time(val: str) -> bool:
    from datetime import datetime
    for fmt in TIME_FORMATS:
        try:
            datetime.strptime(val, fmt)
            return True
        except ValueError:
            continue
    return False


def _check_row(fields: list[dict], row: dict, label: str) -> list[str]:
    out = []
    for f in fields:
        name = f["name"]
        val = str(row.get(name, "")).strip()
        when = f.get("_when")

        if when:
            # 触发值可能有好几个（同一个字段挂在父字段的多个取值下），
            # 匹配规则统一在 wizard_schema.when_active 里，别再各写一份
            if not W.when_active(f, row.get(when[0], "")):
                continue          # 条件没触发，这列不用填

        if not val:
            if f.get("required"):
                out.append(f"{label}：「{name}」必填但为空")
            continue

        opts = f.get("options")
        if opts and f.get("match") != "contains" and f.get("option_match") != "contains":
            multi = f.get("type") in ("checkbox_sync_formily", "multiselect_vue", "multiselect_antd")
            vals = [x.strip() for x in val.replace("，", ",").split(",") if x.strip()] if multi else [val]
            bad = [v for v in vals if v not in opts]
            if bad:
                hint = "、".join(opts[:6]) + ("…" if len(opts) > 6 else "")
                out.append(f"{label}：「{name}」的「{'、'.join(bad)}」不是有效值（可选：{hint}）")

        if f.get("type", "") in ("date_by_label", "date_range_start", "date_range_end"):
            # ⚠ 日期控件不认「9.24 10:00:00」这种写法，而且填不进去也不报错 ——
            #   面板还会一直开着挡住后面的点击，害得后面几个字段也填不上，
            #   最后只报「点了保存并下一步但没跳转」，根本看不出是时间的问题。
            #   所以在这儿就拦住。
            if not _looks_like_time(val):
                out.append(f"{label}：「{name}」的「{val}」不是标准时间格式，"
                           f"要写成 2026-09-24 10:00:00（年-月-日 时:分:秒）")

        if f.get("type") == "number_range_by_label":
            rng = W.parse_range(val)
            if rng is None:
                out.append(f"{label}：「{name}」要填「小-大」两个数字（页面上是 n 天至 m 天），"
                           f"上界填 {W.UNLIMITED} 表示不限。实际是「{val}」")
            elif rng[1] != W.UNLIMITED and rng[0] > rng[1]:
                out.append(f"{label}：「{name}」的「{val}」前面比后面大了，页面要求从小到大")

        mx = f.get("max")
        if mx and len(val) > int(mx):
            out.append(f"{label}：「{name}」{len(val)} 字，超过上限 {mx} 字")

        if f.get("type", "").startswith("upload") and val:
            # 网址交给执行时去下载，这里只拦「既不是网址、本地也没有」的
            if _is_dispimg(val):
                out.append(f"{label}：「{name}」这一格是 WPS 的「嵌入单元格」图片，"
                           f"但没能从文件里抽出图片本体。改成「插入 → 浮于单元格上方」的"
                           f"普通图片，或者填本地路径 / http 网址")
            elif not is_url(val) and not Path(val).exists():
                out.append(f"{label}：「{name}」的图片找不到：{val}"
                           f"（可以填本地路径、直接把图贴进单元格，或者填 http 网址）")
    return out
