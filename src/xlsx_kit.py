"""生成 Excel 模板时那些和业务无关的样式活儿。

## 为什么有这个

六份 template（`template` / `wizard_template` / `pp_template` / `ad_template` /
`dmp_template` / `ab_template`）每一份都自己写了一遍：

  · 三个填色常量 —— 同样的十六进制值声明了 5 遍
  · 表头：加粗、居中、按必填/选填上色、挂批注、算列宽
  · 下面几百行的 number_format（日期列、文本列）
  · 「填写说明」页：三列、A/B 窄 C 宽、C 列自动换行、⚠/■ 开头的加粗
  · 存盘：拼 data/xxx.xlsx、建目录、save、返回字符串

这几样和"这个配置类型有哪些字段"完全无关，是纯粹的 openpyxl 样板。

## 什么该进来、什么不该

**该进来**：换个配置类型行为一模一样的东西（就是上面那些）。
**不该进来**：哪一列必填、列怎么排、说明写什么 —— 那是各家自己的业务，
留在各自的 template 里。

⚠ 判断标准：**加第三个调用方时如果得往这里加一个新参数或新分支，
  那它就不该放进来。** 这个文件里的函数都经得起这一条。

## 颜色

`cond`(FCE4D6) 和 `sku`(FFF2E8) 都是"浅橙"但值不同 —— 分别来自
wizard_template 和 pp_template，**故意保持原样没有统一**：
改颜色会改用户看到的模板长相，那是另一件事，不该混在重构里做。
"""
from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .paths import user_path

AUTHOR = "配置助手"          # 批注署名，六份模板一直都是这个

FILLS = {
    "req": PatternFill("solid", fgColor="FFF2CC"),    # 必填 浅黄
    "opt": PatternFill("solid", fgColor="F2F2F2"),    # 选填 浅灰
    "key": PatternFill("solid", fgColor="DDEBF7"),    # 分组键 / 关联键 浅蓝
    "cond": PatternFill("solid", fgColor="FCE4D6"),   # 条件列（选了别的值才要填）浅橙
    "creative": PatternFill("solid", fgColor="E2EFDA"),  # 创意层 浅绿
    "sku": PatternFill("solid", fgColor="FFF2E8"),    # 按 SKU 展开的列 浅橙
}

# 往下给多少行预设格式。500 是六份模板一直在用的数，
# ⚠ 别改大：openpyxl 会为每个设过格式的单元格真的建一个对象，
#   改成 5000 光建对象就要好几秒，而没人会往模板里粘 5000 行。
DATA_ROWS = 500


def header_cell(ws, col: int, title: str, fill: str = "opt", width=None,
                note: str = "", number_format: str = "", rows: int = DATA_ROWS):
    """写第 col 列的表头，并按需给下面的数据行设格式。

    fill  取 FILLS 里的键；给了不认识的键按 opt 走（不炸模板生成）
    width None = 按标题长度估（和原来六份里的算法一致：14~30 之间）
    note  批注；空字符串就不挂
    """
    cell = ws.cell(row=1, column=col, value=title)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = FILLS.get(fill, FILLS["opt"])
    if note:
        cell.comment = Comment(note, AUTHOR)

    letter = get_column_letter(col)
    ws.column_dimensions[letter].width = (
        width if width is not None else max(14, min(30, len(str(title)) * 2.2)))

    if number_format:
        for r in range(2, rows + 1):
            ws.cell(row=r, column=col).number_format = number_format
    return cell


def freeze_header(ws):
    """冻结第一行。列多的时候往右拉还能看见列名。"""
    ws.freeze_panes = "A2"


def doc_sheet(wb, rows: list, widths=(20, 10, 88), title: str = "填写说明",
              bold_prefixes=("⚠", "■"), bold_exact=()):
    """加一张「填写说明」页。

    rows        [(A列, B列, C列), ...]，直接按顺序往下写
    widths      三列的宽度
    bold_prefixes / bold_exact
                以这些字开头、或正好等于这些字的**单元格**加粗。
                ⚠ 判据作用在每一个单元格上，不只是 A 列 —— 原来六份模板就是
                  这么写的（`for c, v in ...: if str(v).startswith("⚠")`），
                  说明文字里以 ⚠ 开头的那几句也跟着加粗，是有意的。
                ⚠ 各家原来的判据不一样（有的看 ⚠、有的看 ■、有的另外还看
                  「字段」「怎么填」「颜色」），所以这里做成参数，
                  调用方把自己那套传进来 —— 不用回来改这个函数。

    C 列自动换行、全部顶端对齐 —— 说明文字很长，不换行会被截掉看不见。
    """
    doc = wb.create_sheet(title)
    for i, w in enumerate(widths):
        doc.column_dimensions[get_column_letter(i + 1)].width = w

    prefixes = tuple(bold_prefixes)
    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            text = str(v)
            if (prefixes and text.startswith(prefixes)) or text in bold_exact:
                cell.font = Font(bold=True)
    return doc


def save(wb, filename: str) -> str:
    """存到 data/ 下，返回绝对路径的字符串。

    ⚠ 走 user_path 而不是拼相对路径：打包成 exe 之后「当前目录」不是程序目录，
      拼相对路径会把模板生成到用户随便哪个地方（见 src/paths.py 开头）。
    """
    out = user_path("data", filename)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return str(out)
