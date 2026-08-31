"""把用户手上那份 Excel 的表头,和「这个配置类型的模板本该有哪几列」对一遍。

## 为什么

模板列在跑之前就定死(策略中心改了哪些列有效、选了哪几个资源位…),漏一列
`datasource` 是**静默**吞掉的 —— 读表时那一列压根不在 header 里,不报错,跑到页面上
才发现某个字段全空。这里在「载入并检查」之前先把差异摆出来。

## 口径

- 只比表头这一行,不看数据。
- 列名归一:去首尾空白 + 全角空格。大小写不动(中文列名用不上,英文列名 CI 里再说)。
- `missing` = 模板有、你的表没有(要补);`extra` = 你的表多出来的(不会被读,提醒一下)。
- 多 sheet 的(资源位投放 / 价格面板)逐 sheet 比;你的表里找不到对应 sheet 名时,
  按「第一个 sheet」兜底比一次(很多人把模板另存后 sheet 名会变)。
"""
from __future__ import annotations

import logging
from pathlib import Path

log = logging.getLogger(__name__)


def _norm(s) -> str:
    return str(s or "").replace("　", " ").strip()


def _read_headers(data_file: str) -> dict:
    """{sheet名: [列名, ...]}。读不了就返回 {}。"""
    p = Path(data_file)
    if not p.exists():
        return {}
    suffix = p.suffix.lower()
    try:
        if suffix in (".xlsx", ".xlsm"):
            from openpyxl import load_workbook
            wb = load_workbook(p, read_only=True, data_only=True)
            try:
                out = {}
                for name in wb.sheetnames:
                    ws = wb[name]
                    first = next(ws.iter_rows(values_only=True), None) or ()
                    out[name] = [_norm(c) for c in first if _norm(c)]
                return out
            finally:
                wb.close()
        elif suffix == ".csv":
            import csv
            with p.open("r", encoding="utf-8-sig", newline="") as f:
                first = next(csv.reader(f), [])
            return {"(csv)": [_norm(c) for c in first if _norm(c)]}
    except Exception:
        log.warning("读表头失败:%s", data_file, exc_info=True)
    return {}


def compare(expected: dict, data_file: str) -> dict:
    """expected: {sheet名: [列名]}(见 registry.expected_columns)。

    返回 {ok, sheets: {sheet名: {missing:[], extra:[], matched_sheet:str}}}。
    expected 为空(不吃 Excel / 算不出来)时 ok=True、sheets={}。
    """
    if not expected:
        return {"ok": True, "sheets": {}}
    actual = _read_headers(data_file)
    if not actual:
        return {"ok": True, "sheets": {}, "note": "读不到这份表,跳过列对齐"}

    first_actual = next(iter(actual.values()), [])
    out_sheets = {}
    all_ok = True
    for sheet, cols in expected.items():
        want = [_norm(c) for c in cols]
        # sheet 名对得上就用对应的,对不上按第一个 sheet 兜底
        got = actual.get(sheet)
        matched = sheet
        if got is None:
            got, matched = first_actual, (next(iter(actual), "") if len(actual) == 1 else "")
        got_set = set(got)
        want_set = set(want)
        missing = [c for c in want if c not in got_set]
        extra = [c for c in got if c not in want_set]
        if missing:
            all_ok = False
        out_sheets[sheet] = {"missing": missing, "extra": extra, "matched_sheet": matched}
    return {"ok": all_ok, "sheets": out_sheets}


def summarize(diff: dict) -> str:
    """给日志/CLI 用的一句话。没问题返回 ""。"""
    if not diff or diff.get("ok"):
        return ""
    bits = []
    for sheet, d in (diff.get("sheets") or {}).items():
        if d.get("missing"):
            bits.append(f"「{sheet}」缺列：{', '.join(d['missing'])}")
    return "；".join(bits)
