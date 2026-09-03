"""原生商广新（三连竞价推广 auto-v2）的 Excel 模板：一张平表 + 一页填写说明。

⚠ 只服务 mode: ad_v2。

一行 = 一条素材。所有行的 avid / 素材标题 / 封面分别汇进项目的
稿件池 / 标题池 / 封面池（聚合配置）。「内容」列只当人看的备注。
描述在界面「准备」页填一个固定值，不在这张表里。
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from . import ad_prep as P
from . import adv2_data as D
from . import xlsx_kit as X

REQ_FILL = X.FILLS["req"]
OPT_FILL = X.FILLS["opt"]


def build(cfg: dict, prep: dict | None = None, out_name: str | None = None) -> str:
    if prep is None:
        prep = P.load(cfg)

    wb = Workbook()
    ws = wb.active
    ws.title = "素材"

    cols = D.columns(cfg)
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=c["name"])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = REQ_FILL if c.get("required") else OPT_FILL
        ws.column_dimensions[get_column_letter(i)].width = max(14, min(42, len(c["name"]) * 3))
        note = c.get("_note")
        if note:
            cell.comment = Comment(note, "配置助手")
    ws.freeze_panes = "A2"

    _doc_sheet(wb, cfg, prep)

    name = out_name or f"{cfg['name']}_模板"
    return X.save(wb, f"{name}.xlsx")


def _doc_sheet(wb, cfg: dict, prep: dict):
    doc = wb.create_sheet("填写说明")
    for col, w in (("A", 24), ("B", 26), ("C", 82)):
        doc.column_dimensions[col].width = w

    rows = [
        ("配置类型", cfg["name"], cfg.get("description", "")),
        ("", "", ""),
        ("怎么填", "", "① 一行 = 一条素材。"),
        ("", "", "② 新页面素材层是「聚合配置」：整批就一个项目，所有行的 avid 汇进"
                 "稿件池、素材标题汇进标题池、封面汇进封面池，后台自己交叉组合成创意。"),
        ("", "", "③ 重复的 avid / 完全相同的标题会自动去重。"),
        ("", "", "④ 页面上限：稿件 200、标题 50、封面 100。超了跑之前会提示。"),
        ("", "", "⑤ 项目名称、描述、转化目标、出价、预算、投放时间、人群"
                 "这些在界面「准备」页填，不进这张表。"),
        ("", "", "⑥「内容」列只是给你自己看的备注 / 分组标记，程序不读它，留空也行。"),
        ("", "", ""),
        ("颜色", "", "黄=必填　灰=选填"),
        ("封面列", "", "选填。留空 = 那一行不贡献封面；要传就填本地图片路径"
                       "（如 D:\\素材\\a.png），或直接把图片贴进单元格。"
                       "超过 700KB 会自动压到 700KB 以内（只降画质，尺寸不变）。"),
        ("", "", ""),
        ("■ 列说明", "", ""),
    ]
    for c in D.columns(cfg):
        rows.append((f"　{c['name']}", "必填" if c.get("required") else "选填",
                     c.get("_note", "")))

    rows += [("", "", ""), ("■ 准备阶段当前的值", "", "改这些请回界面「准备」页")]
    for name, val in P.summary(cfg, prep):
        rows.append((f"　{name}", val, ""))

    for r, row in enumerate(rows, 1):
        for c, v in enumerate(row, 1):
            cell = doc.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if str(v).startswith("■") or str(v) in ("怎么填", "颜色"):
                cell.font = Font(bold=True)
